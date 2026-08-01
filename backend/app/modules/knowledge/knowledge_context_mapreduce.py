"""资料库大文档 Map-Reduce：分块提炼要点后合并，供报告/计划生成使用。"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
import time
from typing import Any, Optional

from app.core.llm.ai_usage_log import log_ai_usage

from app.core.platform.config import (
    KNOWLEDGE_FULLTEXT_MAX_CHARS,
    KNOWLEDGE_MAPREDUCE_CHUNK_CHARS,
    KNOWLEDGE_MAPREDUCE_CONCURRENCY,
    KNOWLEDGE_MAPREDUCE_ENABLED,
    KNOWLEDGE_MAPREDUCE_MAX_CHUNKS,
)
from app.models.knowledge import AiKnowledgeDocument
from app.modules.ai.ai_prompts import PromptManager
from app.modules.knowledge.knowledge_context import _sections_to_text
from app.modules.knowledge.knowledge_storage import load_document_source
from app.modules.knowledge.report_helpers import extract_json_object

logger = logging.getLogger(__name__)

REPORT_KIND_DIGEST_FOCUS: dict[str, str] = {
    "iteration_report": "侧重迭代测试范围、执行与质量结论相关要点，保留模块与缺陷风险。",
    "functional_report": "侧重功能模块、验收标准、功能范围与缺陷分布相关要点。",
    "performance_report": "侧重性能场景、指标类型、瓶颈与压测范围，勿编造具体数值。",
    "test_plan": "侧重测试目标、范围、里程碑、资源与环境，适合编写测试计划。",
    "test_scheme": "侧重测试类型、策略、组织分工、环境与风险，适合编写测试方案。",
}

_XLSX_MAX_ROWS = 500
_XLSX_MAX_SHEETS = 10


def split_text_chunks(
    text: str,
    *,
    chunk_size: int,
    overlap: int = 400,
    max_chunks: Optional[int] = None,
) -> list[str]:
    """按字数分块，尽量在段落边界切分。"""
    text = (text or "").strip()
    if not text:
        return []
    if len(text) <= chunk_size:
        return [text]

    chunks: list[str] = []
    start = 0
    n = len(text)
    cap = max_chunks if max_chunks is not None else KNOWLEDGE_MAPREDUCE_MAX_CHUNKS
    while start < n and len(chunks) < cap:
        end = min(start + chunk_size, n)
        if end < n:
            cut = text.rfind("\n\n", start + chunk_size // 2, end)
            if cut <= start:
                cut = text.rfind("\n", start + chunk_size // 2, end)
            if cut > start:
                end = cut
        piece = text[start:end].strip()
        if piece:
            chunks.append(piece)
        if end >= n:
            break
        start = max(end - overlap, start + 1)
    return chunks


def _xlsx_to_text(content: bytes) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for name in wb.sheetnames[:_XLSX_MAX_SHEETS]:
            ws = wb[name]
            lines: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= _XLSX_MAX_ROWS:
                    lines.append("…（表格后续行已省略）")
                    break
                cells = [("" if c is None else str(c).strip()) for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))
            if lines:
                parts.append(f"[{name}]\n" + "\n".join(lines))
    finally:
        wb.close()
    return "\n\n".join(parts)


def _xls_to_text(content: bytes) -> str:
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    parts: list[str] = []
    for name in wb.sheet_names()[:_XLSX_MAX_SHEETS]:
        ws = wb.sheet_by_name(name)
        lines: list[str] = []
        for i in range(min(ws.nrows, _XLSX_MAX_ROWS)):
            cells = [("" if c is None else str(c).strip()) for c in ws.row_values(i)]
            if any(cells):
                lines.append(" | ".join(cells))
        if lines:
            parts.append(f"[{name}]\n" + "\n".join(lines))
    return "\n\n".join(parts)


def _csv_to_text(content: bytes) -> str:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines: list[str] = []
    for i, row in enumerate(reader):
        if i >= _XLSX_MAX_ROWS:
            lines.append("…（后续行已省略）")
            break
        lines.append(" | ".join(row))
    return "\n".join(lines)


def extract_document_full_text(doc: AiKnowledgeDocument) -> str:
    """从 sections_json v2 或存储读取尽量完整的正文（Map 阶段用）。"""
    from app.modules.knowledge.knowledge_sections_read import (
        extract_fulltext_from_sections_json,
        is_sections_json_v2,
    )

    text = extract_fulltext_from_sections_json(doc.sections_json)
    if is_sections_json_v2(doc.sections_json) and text:
        return text

    storage = doc.storage or {}
    if not storage:
        return text

    file_name = doc.file_name or ""
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
    try:
        raw = load_document_source(storage, file_name)
    except Exception:
        return text

    try:
        if ext in ("docx", "pdf", "md", "txt"):
            from app.modules.ai.document_loader import load_document

            loaded = load_document(raw, file_name)
            full = (loaded.text or "").strip()
            return full or text
        if ext == "xlsx":
            full = _xlsx_to_text(raw)
            return full or text
        if ext == "xls":
            full = _xls_to_text(raw)
            return full or text
        if ext == "csv":
            full = _csv_to_text(raw)
            return full or text
        if ext in ("doc", "ppt", "pptx"):
            from app.modules.knowledge.knowledge_ingest import parse_document_content

            parsed = parse_document_content(raw, file_name, doc_type=getattr(doc, "doc_type", None))
            full = extract_fulltext_from_sections_json(parsed.get("sections_json"))
            return full or text
    except Exception as ex:
        logger.warning("[mapreduce] extract full text failed doc=%s err=%s", doc.id, ex)
    return text


def should_mapreduce_knowledge(
    knowledge: dict[str, Any],
    settings: Optional[dict[str, Any]] = None,
) -> bool:
    cfg = settings or {}
    enabled = cfg.get("mapreduce_enabled", KNOWLEDGE_MAPREDUCE_ENABLED)
    if not enabled:
        return False
    limit = int(cfg.get("fulltext_max_chars") or KNOWLEDGE_FULLTEXT_MAX_CHARS)
    if knowledge.get("truncated"):
        return True
    raw = int(knowledge.get("raw_total_chars") or 0)
    return raw > limit


def _digest_from_llm_response(raw: str) -> str:
    parsed = extract_json_object(raw)
    for key in ("digest", "merged_digest", "summary"):
        val = parsed.get(key)
        if isinstance(val, str) and val.strip():
            return val.strip()
        if isinstance(val, list):
            lines = [str(x).strip() for x in val if str(x).strip()]
            if lines:
                return "\n".join(f"- {x}" if not x.startswith("-") else x for x in lines)
    return raw.strip()[:8000]


async def _digest_one_chunk(
    *,
    ai_config_id: Optional[int],
    project_name: str,
    iteration_name: str,
    doc_label: str,
    chunk_index: int,
    chunk_total: int,
    chunk_text: str,
    digest_focus: str = "",
    project_id: Optional[int] = None,
    username: str = "",
) -> tuple[str, int]:
    from app.routers.ai.generate import _call_llm, _get_ai_config

    ctx = {
        "project_name": project_name,
        "iteration_name": iteration_name or project_name,
        "doc_label": doc_label,
        "chunk_index": chunk_index,
        "chunk_total": chunk_total,
        "chunk_text": chunk_text,
        "digest_focus": digest_focus or "提炼与测试相关的核心要点。",
    }
    system_prompt, user_prompt = await PromptManager.render("knowledge_chunk_digest", ctx)
    config = await _get_ai_config(ai_config_id, scene="knowledge_chunk_digest")
    t0 = time.time()
    resp = await _call_llm(system_prompt, user_prompt, config, min_timeout=90, max_retries=1)
    digest = _digest_from_llm_response(resp.get("content") or "")
    tokens = int(resp.get("tokens") or 0)
    duration_ms = int((time.time() - t0) * 1000)
    if project_id:
        await log_ai_usage(
            config,
            "knowledge_chunk_digest",
            username=username,
            project_id=project_id,
            tokens_used=tokens,
            duration_ms=duration_ms,
            input_summary=f"{doc_label} · 片段 {chunk_index}/{chunk_total}"[:500],
            output_summary=digest[:500],
        )
    header = f"### {doc_label} · 片段 {chunk_index}/{chunk_total}\n"
    return header + digest, tokens


async def _merge_digests(
    *,
    ai_config_id: Optional[int],
    project_name: str,
    iteration_name: str,
    digests: list[str],
    output_limit: int,
    project_id: Optional[int] = None,
    username: str = "",
) -> tuple[str, int]:
    from app.routers.ai.generate import _call_llm, _get_ai_config

    combined = "\n\n".join(digests).strip()
    if len(combined) <= output_limit:
        return combined, 0

    ctx = {
        "project_name": project_name,
        "iteration_name": iteration_name or project_name,
        "digests": combined[: output_limit * 3],
        "output_char_limit": output_limit,
    }
    system_prompt, user_prompt = await PromptManager.render("knowledge_digest_merge", ctx)
    config = await _get_ai_config(ai_config_id, scene="knowledge_digest_merge")
    t0 = time.time()
    resp = await _call_llm(system_prompt, user_prompt, config, min_timeout=120, max_retries=1)
    merged = _digest_from_llm_response(resp.get("content") or "")
    tokens = int(resp.get("tokens") or 0)
    duration_ms = int((time.time() - t0) * 1000)
    if project_id:
        await log_ai_usage(
            config,
            "knowledge_digest_merge",
            username=username,
            project_id=project_id,
            tokens_used=tokens,
            duration_ms=duration_ms,
            input_summary=f"{project_name} · {iteration_name or project_name} · {len(digests)} 段摘要"[:500],
            output_summary=merged[:500],
        )
    if len(merged) > output_limit:
        merged = merged[:output_limit] + "\n…（合并摘要已截断）"
    return merged, tokens


async def mapreduce_knowledge_context(
    docs: list[AiKnowledgeDocument],
    *,
    ai_config_id: Optional[int],
    project_name: str,
    iteration_name: str = "",
    output_limit: Optional[int] = None,
    settings: Optional[dict[str, Any]] = None,
    report_kind: str = "",
    project_id: Optional[int] = None,
    username: str = "",
) -> dict[str, Any]:
    """对多份资料分块 Map 提炼，Reduce 合并为可送入报告生成的上下文。"""
    cfg = settings or {}
    limit = output_limit or int(cfg.get("fulltext_max_chars") or KNOWLEDGE_FULLTEXT_MAX_CHARS)
    chunk_size = int(cfg.get("mapreduce_chunk_chars") or KNOWLEDGE_MAPREDUCE_CHUNK_CHARS)
    max_chunks = int(cfg.get("mapreduce_max_chunks") or KNOWLEDGE_MAPREDUCE_MAX_CHUNKS)
    concurrency = int(cfg.get("mapreduce_concurrency") or KNOWLEDGE_MAPREDUCE_CONCURRENCY)
    digest_focus = REPORT_KIND_DIGEST_FOCUS.get(report_kind, "") if report_kind else ""

    tasks_meta: list[tuple[str, int, int, str]] = []
    for doc in docs:
        if _is_bug_only_for_stats(doc):
            continue
        label = f"《{doc.title}》({doc.doc_type})"
        full = extract_document_full_text(doc)
        if not full.strip():
            continue
        chunks = split_text_chunks(full, chunk_size=chunk_size, max_chunks=max_chunks)
        for idx, chunk in enumerate(chunks, start=1):
            tasks_meta.append((label, idx, len(chunks), chunk))

    if not tasks_meta:
        return {"text": "", "tokens_used": 0, "chunk_count": 0, "doc_count": 0}

    if len(tasks_meta) > max_chunks:
        tasks_meta = tasks_meta[:max_chunks]
        logger.warning("[mapreduce] capped chunks at %s", max_chunks)

    sem = asyncio.Semaphore(max(1, concurrency))
    tokens_used = 0

    async def _run(item: tuple[str, int, int, str]) -> str:
        nonlocal tokens_used
        label, idx, total, chunk = item
        async with sem:
            digest, tok = await _digest_one_chunk(
                ai_config_id=ai_config_id,
                project_name=project_name,
                iteration_name=iteration_name,
                doc_label=label,
                chunk_index=idx,
                chunk_total=total,
                chunk_text=chunk,
                digest_focus=digest_focus,
                project_id=project_id,
                username=username,
            )
            tokens_used += tok
            return digest

    digests = await asyncio.gather(*[_run(t) for t in tasks_meta])
    merged, merge_tokens = await _merge_digests(
        ai_config_id=ai_config_id,
        project_name=project_name,
        iteration_name=iteration_name,
        digests=list(digests),
        output_limit=limit,
        project_id=project_id,
        username=username,
    )
    tokens_used += merge_tokens

    header = (
        f"【资料分块提炼摘要】共 {len(docs)} 份文档、{len(tasks_meta)} 个片段；"
        f"以下为合并后的核心要点，供报告生成参考。\n\n"
    )
    text = (header + merged).strip()
    if len(text) > limit:
        text = text[:limit] + "\n…（已截断）"

    return {
        "text": text,
        "tokens_used": tokens_used,
        "chunk_count": len(tasks_meta),
        "doc_count": len(docs),
    }


def _is_bug_only_for_stats(doc: AiKnowledgeDocument) -> bool:
    from app.modules.knowledge.knowledge_context import _is_bug_export_candidate

    return _is_bug_export_candidate(doc)


async def maybe_mapreduce_knowledge_context(
    knowledge: dict[str, Any],
    docs: list[AiKnowledgeDocument],
    *,
    ai_config_id: Optional[int],
    project_name: str,
    iteration_name: str = "",
    settings: Optional[dict[str, Any]] = None,
    report_kind: str = "",
    project_id: Optional[int] = None,
    username: str = "",
) -> Optional[dict[str, Any]]:
    """若资料超长则 Map-Reduce；否则返回 None。"""
    if not should_mapreduce_knowledge(knowledge, settings):
        return None
    if not docs:
        return None
    try:
        cfg = settings or {}
        result = await mapreduce_knowledge_context(
            docs,
            ai_config_id=ai_config_id,
            project_name=project_name,
            iteration_name=iteration_name,
            output_limit=int(cfg.get("fulltext_max_chars") or KNOWLEDGE_FULLTEXT_MAX_CHARS),
            settings=settings,
            report_kind=report_kind,
            project_id=project_id,
            username=username,
        )
        if not (result.get("text") or "").strip():
            return None
        return result
    except Exception as ex:
        logger.exception("[mapreduce] failed, fallback to truncated context: %s", ex)
        return None
