"""资料库 v2 表格类文档全量解析（xlsx / xls / csv）"""
from __future__ import annotations

import csv
import io
from typing import Any, Iterator

from app.modules.knowledge.constants import (
    KNOWLEDGE_INGEST_MAX_ROWS_PER_SHEET,
    KNOWLEDGE_INGEST_MAX_SHEETS,
)


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _row_to_dict(headers: list[str], row: list[str], *, row_index: int) -> dict[str, str]:
    item: dict[str, str] = {"__row_index": row_index}
    for idx, header in enumerate(headers):
        if not header:
            continue
        item[header] = row[idx] if idx < len(row) else ""
    return item


def _append_truncation_warning(
    warnings: list[str],
    *,
    sheet_label: str,
    max_rows: int,
    source_row_count: int,
) -> None:
    label = f"「{sheet_label}」" if sheet_label else ""
    warnings.append(
        f"{label}源数据共 {source_row_count} 行，超过上限 {max_rows}，已截断；"
        f"仅保留前 {max_rows} 行"
    )


def _parse_data_rows_stream(
    data_rows: Iterator[tuple[int, list[str]]],
    headers: list[str],
    *,
    max_rows: int,
    warnings: list[str],
    sheet_label: str = "",
) -> tuple[list[dict[str, str]], int, int]:
    """流式解析数据行；返回 (stored_rows, source_row_count, stored_row_count)。"""
    stored: list[dict[str, str]] = []
    source_count = 0
    truncated = False

    for excel_row_idx, cells in data_rows:
        if not any(cells):
            continue
        source_count += 1
        if len(stored) < max_rows:
            stored.append(_row_to_dict(headers, cells, row_index=excel_row_idx))
        else:
            truncated = True

    if truncated:
        _append_truncation_warning(
            warnings,
            sheet_label=sheet_label,
            max_rows=max_rows,
            source_row_count=source_count,
        )
    return stored, source_count, len(stored)


def _sheet_payload(
    *,
    name: str,
    headers: list[str],
    rows: list[dict[str, str]],
    source_row_count: int,
    stored_row_count: int,
) -> dict[str, Any]:
    return {
        "name": name,
        "header_row_index": 0,
        "headers": headers,
        "row_count": source_row_count,
        "stored_row_count": stored_row_count,
        "rows_truncated": source_row_count > stored_row_count,
        "rows": rows,
    }


def _parse_xlsx_sheets(content: bytes, *, max_rows: int, max_sheets: int, warnings: list[str]) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    sheet_names = wb.sheetnames[:max_sheets]
    if len(wb.sheetnames) > max_sheets:
        warnings.append(f"Sheet 数量超过上限 {max_sheets}，仅解析前 {max_sheets} 个")

    for name in sheet_names:
        ws = wb[name]
        row_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration:
            sheets.append(_sheet_payload(name=name, headers=[], rows=[], source_row_count=0, stored_row_count=0))
            continue

        headers = [_cell_str(c) for c in header_row]

        def _data_rows() -> Iterator[tuple[int, list[str]]]:
            excel_row_idx = 2
            for row in row_iter:
                yield excel_row_idx, [_cell_str(c) for c in row]
                excel_row_idx += 1

        rows, source_count, stored_count = _parse_data_rows_stream(
            _data_rows(),
            headers,
            max_rows=max_rows,
            warnings=warnings,
            sheet_label=name,
        )
        sheets.append(
            _sheet_payload(
                name=name,
                headers=headers,
                rows=rows,
                source_row_count=source_count,
                stored_row_count=stored_count,
            )
        )
    wb.close()
    return sheets


def _parse_xls_sheets(content: bytes, *, max_rows: int, max_sheets: int, warnings: list[str]) -> list[dict[str, Any]]:
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    sheets: list[dict[str, Any]] = []
    names = wb.sheet_names()[:max_sheets]
    if len(wb.sheet_names()) > max_sheets:
        warnings.append(f"Sheet 数量超过上限 {max_sheets}，仅解析前 {max_sheets} 个")

    for name in names:
        ws = wb.sheet_by_name(name)
        if ws.nrows <= 0:
            sheets.append(_sheet_payload(name=name, headers=[], rows=[], source_row_count=0, stored_row_count=0))
            continue

        headers = [_cell_str(c) for c in ws.row_values(0)]

        def _data_rows() -> Iterator[tuple[int, list[str]]]:
            for i in range(1, ws.nrows):
                yield i + 1, [_cell_str(c) for c in ws.row_values(i)]

        rows, source_count, stored_count = _parse_data_rows_stream(
            _data_rows(),
            headers,
            max_rows=max_rows,
            warnings=warnings,
            sheet_label=name,
        )
        sheets.append(
            _sheet_payload(
                name=name,
                headers=headers,
                rows=rows,
                source_row_count=source_count,
                stored_row_count=stored_count,
            )
        )
    return sheets


def _parse_csv_sheet(content: bytes, *, max_rows: int, warnings: list[str]) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header_row = next(reader)
    except StopIteration:
        return [
            _sheet_payload(name="Sheet1", headers=[], rows=[], source_row_count=0, stored_row_count=0)
        ]

    headers = [_cell_str(c) for c in header_row]

    def _data_rows() -> Iterator[tuple[int, list[str]]]:
        excel_row_idx = 2
        for row in reader:
            yield excel_row_idx, [_cell_str(c) for c in row]
            excel_row_idx += 1

    rows, source_count, stored_count = _parse_data_rows_stream(
        _data_rows(),
        headers,
        max_rows=max_rows,
        warnings=warnings,
        sheet_label="CSV",
    )
    return [
        _sheet_payload(
            name="Sheet1",
            headers=headers,
            rows=rows,
            source_row_count=source_count,
            stored_row_count=stored_count,
        )
    ]


def _tabular_char_count(sheets: list[dict[str, Any]]) -> int:
    total = 0
    for sh in sheets:
        for row in sh.get("rows") or []:
            if isinstance(row, dict):
                total += sum(len(str(v)) for k, v in row.items() if k != "__row_index")
        for h in sh.get("headers") or []:
            total += len(str(h))
    return total


def parse_tabular_v2(content: bytes, file_name: str, ext: str) -> dict[str, Any]:
    """解析表格文档为 v2 结构（sheets + headers + rows + row_count）。"""
    warnings: list[str] = []
    max_rows = KNOWLEDGE_INGEST_MAX_ROWS_PER_SHEET
    max_sheets = KNOWLEDGE_INGEST_MAX_SHEETS

    if ext == "xlsx":
        sheets = _parse_xlsx_sheets(content, max_rows=max_rows, max_sheets=max_sheets, warnings=warnings)
    elif ext == "xls":
        sheets = _parse_xls_sheets(content, max_rows=max_rows, max_sheets=max_sheets, warnings=warnings)
    elif ext == "csv":
        sheets = _parse_csv_sheet(content, max_rows=max_rows, warnings=warnings)
    else:
        raise ValueError(f"不支持的表格格式: {ext}")

    char_count = _tabular_char_count(sheets)
    return {
        "format": ext,
        "sheets": sheets,
        "char_count": char_count,
        "parse_warnings": warnings,
        "chunk_count": max(len(sheets), 1),
    }
