"""docxtpl / openpyxl 模板渲染与占位符扫描"""
from __future__ import annotations

import io
import re
import zipfile
from typing import Any

from app.modules.knowledge.template_registry import categorize_variables

_JINJA_VAR_RE = re.compile(r"\{\{\s*([a-zA-Z_][\w.]*)\s*")
_JINJA_FOR_LOOP_RE = re.compile(
    r"\{%(?:tr|p)?\s+for\s+([a-zA-Z_]\w*)\s+in\s+([a-zA-Z_][\w.]*)"
)
_JINJA_SET_RE = re.compile(r"\{%\s*set\s+([a-zA-Z_][\w.]*)\s*=")
_JINJA_MACRO_RE = re.compile(r"\{%\s*macro\s+([a-zA-Z_][\w.]*)")
# openpyxl 单元格 {{ var }}
_CELL_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z_][\w.]*)\s*\}\}")


def _root_name(token: str) -> str:
    return (token or "").split(".")[0]


def _scan_jinja_text(text: str) -> tuple[set[str], set[str]]:
    """从 XML/文本中提取模板变量与 for 循环迭代器名。"""
    found: set[str] = set()
    loop_iterators: set[str] = set()

    for m in _JINJA_FOR_LOOP_RE.finditer(text):
        loop_iterators.add(_root_name(m.group(1)))
        found.add(_root_name(m.group(2)))

    for m in _JINJA_VAR_RE.finditer(text):
        found.add(_root_name(m.group(1)))

    for m in _JINJA_SET_RE.finditer(text):
        found.add(_root_name(m.group(1)))

    for m in _JINJA_MACRO_RE.finditer(text):
        found.add(_root_name(m.group(1)))

    for m in _CELL_PLACEHOLDER_RE.finditer(text):
        found.add(_root_name(m.group(1)))

    # {{ row.field }} 中的 row 若仅为循环变量，不参与「未知占位符」校验
    found -= loop_iterators
    return found, loop_iterators


def _extract_docx_placeholders(content: bytes) -> tuple[set[str], set[str]]:
    found: set[str] = set()
    loop_iterators: set[str] = set()
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for name in zf.namelist():
                if not name.startswith("word/") or not name.endswith(".xml"):
                    continue
                try:
                    text = zf.read(name).decode("utf-8", errors="ignore")
                except Exception:
                    continue
                part_found, part_iters = _scan_jinja_text(text)
                found |= part_found
                loop_iterators |= part_iters
    except zipfile.BadZipFile:
        pass
    return found, loop_iterators


def _extract_xlsx_placeholders(content: bytes) -> set[str]:
    import openpyxl

    found: set[str] = set()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=200, max_col=30):
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                for m in _CELL_PLACEHOLDER_RE.finditer(val):
                    found.add(m.group(1).split(".")[0])
    wb.close()
    return found


def scan_template_placeholders(content: bytes, file_name: str) -> dict[str, Any]:
    ext = (file_name or "").rsplit(".", 1)[-1].lower() if "." in (file_name or "") else ""
    loop_iterators: set[str] = set()
    if ext == "docx":
        found, loop_iterators = _extract_docx_placeholders(content)
    elif ext in ("xlsx", "xls"):
        found = _extract_xlsx_placeholders(content)
    else:
        found = set()
    cats = categorize_variables(found)
    return {
        "placeholders": sorted(found),
        "known": cats["known"],
        "unknown": cats["unknown"],
        "loop_iterators": sorted(loop_iterators),
        "has_unknown": bool(cats["unknown"]),
    }


def render_docx_template(template_bytes: bytes, context: dict[str, Any]) -> bytes:
    from docxtpl import DocxTemplate

    tpl = DocxTemplate(io.BytesIO(template_bytes))
    tpl.render(context)
    out = io.BytesIO()
    tpl.save(out)
    return out.getvalue()


def render_xlsx_template(template_bytes: bytes, context: dict[str, Any]) -> bytes:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or "{{" not in val:
                    continue
                new_val = val
                for m in _CELL_PLACEHOLDER_RE.finditer(val):
                    key = m.group(1).split(".")[0]
                    if key in context:
                        rep = str(context[key]) if context[key] is not None else ""
                        new_val = new_val.replace(m.group(0), rep)
                cell.value = new_val
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def render_template(template_bytes: bytes, file_name: str, context: dict[str, Any]) -> bytes:
    ext = (file_name or "").rsplit(".", 1)[-1].lower() if "." in (file_name or "") else "docx"
    if ext in ("xlsx", "xls"):
        return render_xlsx_template(template_bytes, context)
    return render_docx_template(template_bytes, context)
