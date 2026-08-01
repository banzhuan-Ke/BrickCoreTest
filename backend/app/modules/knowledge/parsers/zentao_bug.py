"""禅道 Bug 导出 xlsx/csv 通用解析"""
from __future__ import annotations

import csv
import io
import re
from collections import Counter
from typing import Any, Optional

# 表头别名 → 标准字段
_BUG_HEADER_ALIASES: dict[str, str] = {
    "bug编号": "bug_id",
    "编号": "bug_id",
    "缺陷编号": "bug_id",
    "缺陷id": "bug_id",
    "id": "bug_id",
    "bug id": "bug_id",
    "bug标题": "title",
    "标题": "title",
    "缺陷标题": "title",
    "缺陷名称": "title",
    "所属模块": "module",
    "模块": "module",
    "严重程度": "severity",
    "优先级": "priority",
    "bug状态": "status",
    "状态": "status",
    "解决方案": "resolution",
    "解决者": "resolved_by",
    "解决人": "resolved_by",
    "由谁解决": "resolved_by",
    "resolvedby": "resolved_by",
    "指派给": "assigned_to",
    "负责人": "assigned_to",
    "归属人": "assigned_to",
    "创建者": "opened_by",
    "创建人": "opened_by",
    "由谁创建": "opened_by",
    "创建日期": "opened_date",
    "关闭日期": "closed_date",
    "关闭者": "closed_by",
    "解决日期": "closed_date",
    "所属产品": "product",
    "所属项目": "project",
    "所属执行": "execution",
    "反馈者": "feedback",
    "Bug类型": "bug_type",
    "bug归属人": "bug_owner",
    "BUG归属人": "bug_owner",
    "測試區域": "test_zone",
    "测试区域": "test_zone",
    "BUG分类": "bug_class",
    "bug分类": "bug_class",
    "问题原因（确认有找到根本原因）": "root_cause",
    "根因分类": "root_cause_category",
}


def _norm_header(cell: Any) -> str:
    return re.sub(r"\s+", "", str(cell or "").strip().lower())


def _map_header(cell: Any) -> Optional[str]:
    raw = str(cell or "").strip()
    if not raw:
        return None
    if raw in ("測試區域", "BUG分类", "BUG归属人"):
        return {"測試區域": "test_zone", "BUG分类": "bug_class", "BUG归属人": "bug_owner"}[raw]
    key = _norm_header(cell)
    if not key:
        return None
    if key in _BUG_HEADER_ALIASES:
        return _BUG_HEADER_ALIASES[key]
    for alias, field in _BUG_HEADER_ALIASES.items():
        if alias in key or key in alias:
            return field
    return None


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_rows_from_xlsx(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["Bug"] if "Bug" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        mapped = _map_header(cell)
        if mapped:
            col_map[idx] = mapped

    bugs: list[dict[str, str]] = []
    for row in rows_iter:
        if not any(row):
            continue
        item: dict[str, str] = {}
        for idx, field in col_map.items():
            if idx < len(row):
                item[field] = _cell_str(row[idx])
        if item.get("bug_id") or item.get("title"):
            bugs.append(item)
    wb.close()
    # Bug 编号去重
    seen: set[str] = set()
    deduped: list[dict[str, str]] = []
    for item in bugs:
        key = item.get("bug_id") or ""
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        deduped.append(item)
    headers = list(col_map.values())
    return headers, deduped


def _parse_rows_from_csv(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header_row = next(reader)
    except StopIteration:
        return [], []

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        mapped = _map_header(cell)
        if mapped:
            col_map[idx] = mapped

    bugs: list[dict[str, str]] = []
    for row in reader:
        if not any(row):
            continue
        item: dict[str, str] = {}
        for idx, field in col_map.items():
            if idx < len(row):
                item[field] = _cell_str(row[idx])
        if item.get("bug_id") or item.get("title"):
            bugs.append(item)
    return list(col_map.values()), bugs


def parse_zentao_bug_export(content: bytes, file_name: str = "") -> dict[str, Any]:
    """解析禅道 Bug 导出，返回 bugs 列表与统计摘要。"""
    ext = (file_name or "").rsplit(".", 1)[-1].lower() if "." in (file_name or "") else ""
    if ext == "csv":
        headers, bugs = _parse_rows_from_csv(content)
    else:
        headers, bugs = _parse_rows_from_xlsx(content)

    severity_counter = Counter(b.get("severity") or "未知" for b in bugs)
    status_counter = Counter(b.get("status") or "未知" for b in bugs)
    module_counter = Counter(b.get("module") or "未分类" for b in bugs)
    resolution_counter = Counter(b.get("resolution") or "未填写" for b in bugs if b.get("resolution"))
    opened_by_counter = Counter(b.get("opened_by") or "未知" for b in bugs if b.get("opened_by"))
    assigned_counter = Counter(b.get("assigned_to") or "未指派" for b in bugs if b.get("assigned_to"))
    resolved_by_counter = Counter(b.get("resolved_by") or "未知" for b in bugs if b.get("resolved_by"))
    priority_counter = Counter(b.get("priority") or "未知" for b in bugs if b.get("priority"))

    closed_hints = ("关闭", "已解决", "resolved", "closed", "已关闭")
    open_count = sum(1 for b in bugs if not any(h in (b.get("status") or "").lower() for h in closed_hints))
    closed_count = len(bugs) - open_count

    summary_lines = [
        f"共 {len(bugs)} 条 Bug",
    ]
    if severity_counter:
        sev = "；".join(f"{k}:{v}" for k, v in severity_counter.most_common(8))
        summary_lines.append(f"严重程度：{sev}")
    if status_counter:
        st = "；".join(f"{k}:{v}" for k, v in status_counter.most_common(8))
        summary_lines.append(f"状态：{st}")
    if resolution_counter:
        res = "；".join(f"{k}:{v}" for k, v in resolution_counter.most_common(8))
        summary_lines.append(f"解决方案：{res}")
    if opened_by_counter:
        ob = "；".join(f"{k}:{v}" for k, v in opened_by_counter.most_common(6))
        summary_lines.append(f"创建人：{ob}")
    if assigned_counter:
        asg = "；".join(f"{k}:{v}" for k, v in assigned_counter.most_common(6))
        summary_lines.append(f"归属/指派：{asg}")
    if resolved_by_counter:
        rb = "；".join(f"{k}:{v}" for k, v in resolved_by_counter.most_common(6))
        summary_lines.append(f"解决人：{rb}")

    return {
        "headers_found": headers,
        "bugs": bugs,
        "total": len(bugs),
        "by_severity": dict(severity_counter),
        "by_status": dict(status_counter),
        "by_module": dict(module_counter.most_common(20)),
        "by_resolution": dict(resolution_counter),
        "by_opened_by": dict(opened_by_counter),
        "by_assigned_to": dict(assigned_counter),
        "by_resolved_by": dict(resolved_by_counter),
        "by_priority": dict(priority_counter),
        "open_count": open_count,
        "closed_count": closed_count,
        "summary_text": "\n".join(summary_lines),
    }
