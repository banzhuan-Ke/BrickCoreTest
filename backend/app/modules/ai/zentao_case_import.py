"""
禅道用例 XLSX 导入解析
"""
import io
import uuid
from typing import Any, Optional

from app.core.case.case_steps import align_steps_expects, split_numbered_lines
from app.modules.ai.zentao_bindings import resolve_stage
from app.modules.ai.zentao_case_export import ZENTAO_DEFAULT_COLUMNS
from app.modules.ai.zentao_case_types import normalize_zentao_case_type

# 表头别名 → 标准列名
_HEADER_ALIASES = {
    "所属产品": "所属产品",
    "产品": "所属产品",
    "所属模块": "所属模块",
    "模块": "所属模块",
    "相关研发需求": "相关研发需求",
    "研发需求": "相关研发需求",
    "用例标题": "用例标题",
    "标题": "用例标题",
    "前置条件": "前置条件",
    "步骤": "步骤",
    "预期": "预期",
    "优先级": "优先级",
    "用例类型": "用例类型",
    "类型": "用例类型",
    "适用阶段": "适用阶段",
    "阶段": "适用阶段",
}

# 禅道用例 ID 列（可选，导入时写入 zentao_case_id）
_ZENTAO_ID_HEADERS = frozenset({
    "用例编号", "编号", "用例ID", "用例 id", "ID", "id", "禅道ID", "禅道id",
})


def _map_zentao_id_header(cell: str) -> bool:
    s = (cell or "").strip()
    return s in _ZENTAO_ID_HEADERS


def _map_header(cell: str) -> Optional[str]:
    s = (cell or "").strip()
    if not s:
        return None
    if s in ZENTAO_DEFAULT_COLUMNS:
        return s
    return _HEADER_ALIASES.get(s)


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_xlsx_rows(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    col_map: dict[int, str] = {}
    zentao_id_col: Optional[int] = None
    for idx, cell in enumerate(header_row):
        h = _cell_str(cell)
        if _map_zentao_id_header(h):
            zentao_id_col = idx
            continue
        mapped = _map_header(h)
        if mapped:
            col_map[idx] = mapped

    if "用例标题" not in col_map.values():
        raise ValueError("未识别到「用例标题」列，请使用平台导出的禅道模板")

    parsed: list[dict[str, str]] = []
    for row in rows_iter:
        if not row or all(_cell_str(c) == "" for c in row):
            continue
        item = {col: "" for col in ZENTAO_DEFAULT_COLUMNS}
        for idx, col_name in col_map.items():
            if idx < len(row):
                item[col_name] = _cell_str(row[idx])
        if zentao_id_col is not None and zentao_id_col < len(row):
            item["_zentao_case_id"] = _cell_str(row[zentao_id_col])
        if item.get("用例标题"):
            parsed.append(item)
    return list(col_map.values()), parsed


def _group_rows_by_title(raw_rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    groups: dict[str, list[dict[str, str]]] = {}
    for row in raw_rows:
        title = row.get("用例标题", "").strip()
        if not title:
            continue
        groups.setdefault(title, []).append(row)
    return groups


def parse_zentao_xlsx_to_cases(
    content: bytes,
    export_defaults: Optional[dict] = None,
) -> tuple[list[dict[str, Any]], list[dict], list[dict], str]:
    """
  解析 XLSX → 用例 dict 列表（可写入 AiFunctionalCase）
  返回: cases, failed_rows, warnings, import_batch
    """
    defaults = export_defaults or {}
    import_batch = str(uuid.uuid4())
    failed_rows: list[dict] = []
    warnings: list[dict] = []

    try:
        _, raw_rows = _parse_xlsx_rows(content)
    except Exception as e:
        raise ValueError(f"XLSX 解析失败: {e}") from e

    if not raw_rows:
        raise ValueError("文件中没有有效数据行")

    grouped = _group_rows_by_title(raw_rows)
    cases: list[dict[str, Any]] = []

    for title, rows in grouped.items():
        first = rows[0]
        pairs: list[dict] = []
        for r in rows:
            step_text = r.get("步骤", "")
            expect_text = r.get("预期", "")
            if step_text or expect_text:
                step_lines = split_numbered_lines(step_text) if step_text else [""]
                expect_lines = split_numbered_lines(expect_text) if expect_text else [""]
                n = max(len(step_lines), len(expect_lines), 1)
                for i in range(n):
                    s = step_lines[i] if i < len(step_lines) else ""
                    e = expect_lines[i] if i < len(expect_lines) else ""
                    if s or e:
                        pairs.append({"step": s, "expect": e})
            elif len(rows) > 1:
                warnings.append({
                    "title": title,
                    "reason": "存在无步骤/预期的空行，已跳过",
                })

        steps = align_steps_expects(pairs if pairs else [{"step": "执行操作", "expect": "符合预期"}])
        priority = str(first.get("优先级") or "2").strip() or "2"
        stage = (first.get("适用阶段") or "").strip() or resolve_stage(priority, defaults)

        zentao_id = ""
        for r in rows:
            zid = (r.get("_zentao_case_id") or "").strip()
            if zid:
                zentao_id = zid
                break

        cases.append({
            "product": (first.get("所属产品") or defaults.get("product") or "").strip(),
            "module": (first.get("所属模块") or defaults.get("module") or "").strip(),
            "related_story": (first.get("相关研发需求") or defaults.get("related_story") or "").strip(),
            "title": title[:500],
            "zentao_case_id": zentao_id or None,
            "precondition": first.get("前置条件", ""),
            "steps": steps,
            "priority": priority,
            "type": normalize_zentao_case_type(first.get("用例类型")),
            "stage": stage,
            "keywords": "",
            "source_import_batch": import_batch,
        })

    return cases, failed_rows, warnings, import_batch
