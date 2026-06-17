"""
禅道用例导入 Excel 导出（默认列模板与单元格格式）
"""
import io
import re
from typing import Any, Optional

from app.core.zentao_case_types import normalize_zentao_case_type

# 禅道导入默认列（与平台导出模板一致）
ZENTAO_DEFAULT_COLUMNS = [
    "所属产品",
    "所属模块",
    "相关研发需求",
    "用例标题",
    "前置条件",
    "步骤",
    "预期",
    "优先级",
    "用例类型",
    "适用阶段",
]

# 功能用例库导出列（禅道用例编号放首列，便于回导禅道）
ZENTAO_EXPORT_COLUMNS = [
    "用例编号",
    "所属产品",
    "所属模块",
    "相关研发需求",
    "用例标题",
    "前置条件",
    "步骤",
    "预期",
    "优先级",
    "用例类型",
    "适用阶段",
]

DEFAULT_EXPORT_DEFAULTS = {
    "product": "",
    "module": "",
    "related_story": "",
    "default_stage": "系统测试阶段",
}

STAGE_OPTIONS = ("冒烟测试阶段", "系统测试阶段", "集成测试阶段", "回归测试阶段")


def format_numbered_cell(steps: list, field: str) -> str:
    """将步骤/预期列表格式化为「1.xxx\\n2.xxx」单单元格文本"""
    lines: list[str] = []
    idx = 0
    for st in steps:
        if isinstance(st, dict):
            text = (st.get(field) or "").strip()
        else:
            text = str(st).strip() if st else ""
        if not text:
            continue
        idx += 1
        if re.match(r"^\d+[.、．]\s*", text):
            lines.append(text)
        else:
            lines.append(f"{idx}.{text}")
    return "\n".join(lines)


def case_to_export_row(case: dict, defaults: Optional[dict] = None) -> dict[str, Any]:
    """单条用例 → 导出行，步骤/预期合并为单单元格（1.xxx\\n2.xxx）"""
    defaults = {**DEFAULT_EXPORT_DEFAULTS, **(defaults or {})}
    from app.core.case_steps import align_steps_expects, normalize_corner_quotes

    steps = align_steps_expects(case.get("steps") or [])
    steps_text = format_numbered_cell(steps, "step")
    expects_text = format_numbered_cell(steps, "expect")

    from app.core.zentao_bindings import resolve_stage

    priority = str(case.get("priority") or "2")
    stage = (case.get("stage") or "").strip() or resolve_stage(priority, defaults)
    extra = case.get("extra") if isinstance(case.get("extra"), dict) else {}
    zentao_module = (
        (case.get("zentao_module") or "").strip()
        or (extra.get("zentao_module") or "").strip()
        or (defaults.get("module") or "").strip()
    )

    zentao_id = (
        (case.get("zentao_case_id") or "").strip()
        or (case.get("zentao_case_id_display") or "").strip()
    )

    return {
        "所属产品": (case.get("product") or defaults.get("product") or "").strip(),
        "所属模块": zentao_module,
        # 批量生成时各批次 related_story 已写入用例，导出须以用例为准
        "相关研发需求": (
            case.get("related_story") or defaults.get("related_story") or ""
        ).strip(),
        "用例标题": (case.get("title") or "").strip(),
        "用例编号": zentao_id,
        "前置条件": normalize_corner_quotes((case.get("precondition") or "").strip()),
        "步骤": steps_text,
        "预期": expects_text,
        "优先级": priority,
        "用例类型": normalize_zentao_case_type(case.get("type")),
        "适用阶段": stage,
    }


def build_xlsx_bytes(
    rows: list[dict],
    columns: Optional[list[str]] = None,
    *,
    lightweight: bool = False,
) -> bytes:
    """生成 xlsx 二进制内容；lightweight 用于大批量导出，跳过逐行样式以减轻 CPU"""
    cols = columns or ZENTAO_DEFAULT_COLUMNS
    if lightweight:
        from openpyxl import Workbook

        wb = Workbook(write_only=True)
        ws = wb.create_sheet("用例")
        ws.append(cols)
        for row_data in rows:
            ws.append([row_data.get(c, "") for c in cols])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cols = columns or ZENTAO_DEFAULT_COLUMNS
    wb = Workbook()
    ws = wb.active
    ws.title = "用例"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.append(cols)
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_data in rows:
        ws.append([row_data.get(c, "") for c in cols])

    # 步骤、预期列自动换行
    step_col = cols.index("步骤") + 1 if "步骤" in cols else None
    expect_col = cols.index("预期") + 1 if "预期" in cols else None
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 72
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).alignment = wrap
        if step_col:
            ws.cell(row=r, column=step_col).alignment = wrap
        if expect_col:
            ws.cell(row=r, column=expect_col).alignment = wrap

    widths = {
        "所属产品": 22,
        "所属模块": 24,
        "相关研发需求": 36,
        "用例标题": 32,
        "用例编号": 12,
        "前置条件": 28,
        "步骤": 40,
        "预期": 40,
        "优先级": 8,
        "用例类型": 12,
        "适用阶段": 14,
    }
    for i, name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 16)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
