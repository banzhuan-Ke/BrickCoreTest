"""
知识库问答准确性评测
"""
import asyncio
import io
import json
import logging
from typing import Any, Literal, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.core.auth import is_authenticated, require_permissions, resolve_current_username
from app.core.permissions import AI_TEST_EXECUTE, AI_TEST_VIEW
from app.core.qa_eval_target_client import QA_QUESTION_TYPES
from app.core.qa_eval_compare import CompareGroupSpec, build_compare_report
from app.core.qa_eval_report import (
    QA_IMPORT_TEMPLATE_COLUMNS,
    build_import_template_rows,
    build_merged_statistics_report,
    build_run_statistics_report,
)
from app.core.qa_eval_service import (
    CASE_SCOPE_ALL,
    CASE_SCOPE_RANGE,
    CASE_SCOPE_RETRY_FAILED,
    DEFAULT_BATCH_CHUNK_SIZE,
    MAX_CASES_PER_RUN,
    QA_FETCH_EXPORT_COLUMNS,
    RUN_MODE_AUTO,
    RUN_MODE_FETCH_ONLY,
    RUN_MODE_JUDGE_ONLY,
    RUN_MODE_LABELS,
    build_merged_export_rows,
    build_qa_cases_export_rows,
    build_qa_fetch_export_rows,
    case_to_dict,
    create_batch_runs,
    delete_eval_run,
    has_active_run,
    analyze_seq_numbers,
    list_ordered_cases,
    regenerate_result,
    resolve_cases_for_run,
    result_to_dict,
    run_qa_eval_background,
    run_to_dict,
    set_to_dict,
    target_to_dict,
    test_target_api,
)
from app.core.zentao_case_export import build_xlsx_bytes
from app.models.ai import (
    AiQaEvalCase,
    AiQaEvalResult,
    AiQaEvalRun,
    AiQaEvalSet,
    AiQaEvalTarget,
)
from app.schemas.ai import StandardResponse

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/qa-eval", tags=["问答准确性评测"])

MAX_IMPORT_BYTES = 5 * 1024 * 1024
CASE_TYPES = ("事实", "流程", "拒答", "边界")


async def _qa_xlsx_bytes(rows: list[dict]) -> bytes:
    """大批量问答导出：轻量写表 + 线程池，避免阻塞事件循环"""
    return await asyncio.to_thread(
        build_xlsx_bytes,
        rows,
        QA_FETCH_EXPORT_COLUMNS,
        lightweight=True,
    )


def _resolve_project_id(user_info: dict, project_id: Optional[int] = None) -> int:
    pid = project_id or user_info.get("project_id") or user_info.get("current_project_id")
    if not pid:
        raise HTTPException(status_code=400, detail="请先在顶部导航栏选择项目")
    return int(pid)


async def _get_set(set_id: int, project_id: int) -> AiQaEvalSet:
    s = await AiQaEvalSet.get_or_none(id=set_id, project_id=project_id, is_del=False)
    if not s:
        raise HTTPException(status_code=404, detail="评测集不存在")
    return s


class SetBody(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    description: Optional[str] = None


class CaseBody(BaseModel):
    question: str = Field(..., min_length=1, description="测试问题")
    expected_answer: str = Field(..., min_length=1, description="标准答案/参考答案，传给 LLM 的 ground_truth")
    preset_answer: Optional[str] = Field(None, description="预置实际回答，仅评判模式或外部已跑批时使用")
    expected_points: list[str] = Field(default_factory=list, description="可选补充要点")
    seq_no: Optional[int] = Field(None, description="Excel 序号")
    chat_path: list[Any] = Field(default_factory=list, description="问答目录 ID 列表")
    multi_turn: bool = False
    scenario_type: str = ""
    source_file: str = ""
    file_type: str = ""
    category: str = ""
    case_type: str = "事实"
    sort_order: int = 0


class TargetBody(BaseModel):
    name: str = Field(..., min_length=1, max_length=200)
    config: dict[str, Any] = Field(default_factory=dict)


class TargetTestBody(BaseModel):
    question: str = Field(..., min_length=1, description="调试用的测试问题")
    config: Optional[dict[str, Any]] = Field(None, description="未保存时传入完整 config")
    extra: Optional[dict[str, Any]] = Field(
        None,
        description="扩展变量：chatPath、sessionId、historyFlag 等",
    )


class RunBody(BaseModel):
    run_name: str = Field("", max_length=100, description="评测名称，便于执行记录识别")
    run_mode: Literal["auto", "judge_only", "fetch_only"] = RUN_MODE_AUTO
    target_id: Optional[int] = None
    judge_config_id: Optional[int] = None
    case_scope: Literal["all", "range", "retry_failed"] = CASE_SCOPE_ALL
    range_start: Optional[int] = Field(None, ge=1)
    range_end: Optional[int] = Field(None, ge=1)
    retry_source_run_id: Optional[int] = None
    request_interval_ms: int = Field(0, ge=0, le=120_000)


class BatchRunBody(RunBody):
    chunk_size: int = Field(DEFAULT_BATCH_CHUNK_SIZE, ge=1, le=MAX_CASES_PER_RUN)


class MergeExportBody(BaseModel):
    run_ids: Optional[list[int]] = None
    batch_group_id: Optional[str] = None


class CompareGroupItem(BaseModel):
    label: str = Field(..., min_length=1, max_length=100)
    run_ids: Optional[list[int]] = None
    batch_group_id: Optional[str] = None


class CompareReportBody(BaseModel):
    groups: list[CompareGroupItem] = Field(..., min_length=2, max_length=10)


class ResultReviewBody(BaseModel):
    manual_status: Literal["pending", "approved", "rejected"] = "approved"
    manual_comment: Optional[str] = None


class BulkReviewBody(BaseModel):
    result_ids: list[int] = Field(..., min_length=1)
    manual_status: Literal["pending", "approved", "rejected"] = "approved"
    manual_comment: Optional[str] = None


def _normalize_case_input(
    question: str,
    expected_answer: Optional[str],
    expected_points: list[str],
) -> tuple[str, list[str], str]:
    q = (question or "").strip()
    ans = (expected_answer or "").strip()
    points = [p.strip() for p in (expected_points or []) if p and str(p).strip()]
    if not q:
        raise HTTPException(status_code=400, detail="问题不能为空")
    if not ans:
        raise HTTPException(status_code=400, detail="标准答案不能为空")
    return q, points, ans


def _parse_chat_path_cell(val: Any) -> list[str]:
    if val is None:
        return []
    if isinstance(val, list):
        return [str(x).strip() for x in val if str(x).strip()]
    text = str(val).strip()
    if not text:
        return []
    if text.startswith("["):
        try:
            data = json.loads(text)
            if isinstance(data, list):
                return [str(x).strip() for x in data if str(x).strip()]
        except json.JSONDecodeError:
            pass
    for sep in (";", "；", ",", "，", "/"):
        if sep in text:
            return [p.strip() for p in text.split(sep) if p.strip()]
    return [text]


def _parse_bool_cell(val: Any) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    text = str(val).strip().lower()
    return text in ("1", "true", "yes", "y", "是", "开启", "开", "多轮")


def _parse_int_cell(val: Any) -> Optional[int]:
    if val is None or str(val).strip() == "":
        return None
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _case_fields_from_body(body: CaseBody) -> dict[str, Any]:
    chat_path = [str(x).strip() for x in (body.chat_path or []) if str(x).strip()]
    return {
        "seq_no": body.seq_no,
        "chat_path": chat_path,
        "multi_turn": bool(body.multi_turn),
        "scenario_type": (body.scenario_type or "").strip() or None,
        "source_file": (body.source_file or "").strip() or None,
        "file_type": (body.file_type or "").strip() or None,
    }


def _normalize_run_name(name: str, *, set_name: str = "") -> str:
    text = (name or "").strip()
    if text:
        return text[:100]
    return (set_name or "").strip()[:100]


def _build_run_extra(
    *,
    user_info: dict,
    username: str,
    body: RunBody,
    set_name: str = "",
) -> dict[str, Any]:
    extra: dict[str, Any] = {
        "user_id": user_info.get("id"),
        "username": username,
        "run_name": _normalize_run_name(body.run_name, set_name=set_name),
        "judge_config_id": body.judge_config_id,
        "run_mode": body.run_mode,
        "case_scope": body.case_scope,
        "request_interval_ms": body.request_interval_ms,
        "done_count": 0,
        "progress_percent": 0,
        "current_question": "",
    }
    if body.case_scope == CASE_SCOPE_RANGE:
        extra["range_start"] = body.range_start
        extra["range_end"] = body.range_end
    if body.case_scope == CASE_SCOPE_RETRY_FAILED:
        extra["retry_source_run_id"] = body.retry_source_run_id
    return extra


async def _validate_run_body(
    set_id: int,
    pid: int,
    body: RunBody,
) -> tuple[Optional[AiQaEvalTarget], int]:
    run_mode = body.run_mode or RUN_MODE_AUTO
    if run_mode not in RUN_MODE_LABELS:
        raise HTTPException(status_code=400, detail="无效的 run_mode")

    target: Optional[AiQaEvalTarget] = None
    target_id: Optional[int] = body.target_id
    if run_mode in (RUN_MODE_AUTO, RUN_MODE_FETCH_ONLY):
        if not target_id:
            raise HTTPException(status_code=400, detail="该模式需选择被测 API")
        target = await AiQaEvalTarget.get_or_none(id=target_id, project_id=pid, is_del=False)
        if not target:
            raise HTTPException(status_code=404, detail="被测 API 配置不存在")
    elif run_mode == RUN_MODE_JUDGE_ONLY:
        target_id = body.target_id

    if body.case_scope == CASE_SCOPE_RANGE:
        if body.range_start is None or body.range_end is None:
            raise HTTPException(status_code=400, detail="范围跑批需填写起始与结束序号")
        if body.range_start > body.range_end:
            raise HTTPException(status_code=400, detail="起始序号不能大于结束序号")

    if body.case_scope == CASE_SCOPE_RETRY_FAILED:
        if not body.retry_source_run_id:
            raise HTTPException(status_code=400, detail="重跑失败需指定来源跑批 ID")
        src = await AiQaEvalRun.get_or_none(
            id=body.retry_source_run_id, set_id=set_id, project_id=pid
        )
        if not src:
            raise HTTPException(status_code=404, detail="来源跑批记录不存在")

    probe = AiQaEvalRun(
        set_id=set_id,
        project_id=pid,
        target_id=target_id,
        extra=_build_run_extra(user_info={}, username="", body=body),
    )
    cases = await resolve_cases_for_run(probe)
    if not cases:
        raise HTTPException(status_code=400, detail="没有符合条件的用例")
    if len(cases) > MAX_CASES_PER_RUN:
        raise HTTPException(
            status_code=400,
            detail=f"本次用例数 {len(cases)} 超过上限 {MAX_CASES_PER_RUN}，请缩小范围或使用自动分批",
        )

    if run_mode == RUN_MODE_JUDGE_ONLY:
        missing = sum(1 for c in cases if not (c.preset_answer or "").strip())
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"{missing} 条用例缺少「实际回答」，请 Excel 导入该列或在用例表单中填写",
            )

    return target, len(cases)


def _parse_points_cell(val: Any) -> list[str]:
    if val is None:
        return []
    if isinstance(val, list):
        return [str(x).strip() for x in val if str(x).strip()]
    text = str(val).strip()
    if not text:
        return []
    for sep in ("；", ";", "\n", "|"):
        if sep in text:
            return [p.strip() for p in text.split(sep) if p.strip()]
    return [text]


@router.get("/sets", response_model=StandardResponse)
async def list_sets(
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    pid = _resolve_project_id(user_info, project_id)
    sets = await AiQaEvalSet.filter(project_id=pid, is_del=False).order_by("-id")
    out = []
    for s in sets:
        cnt = await AiQaEvalCase.filter(set_id=s.id, is_del=False).count()
        seq_list = await AiQaEvalCase.filter(set_id=s.id, is_del=False).values_list("seq_no", flat=True)
        seq_meta = analyze_seq_numbers(list(seq_list))
        out.append(
            set_to_dict(
                s,
                case_count=cnt,
                max_seq_no=seq_meta.get("max_seq_no"),
            )
        )
    return StandardResponse(data=out)


@router.post("/sets", response_model=StandardResponse)
async def create_set(
    body: SetBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    username = await resolve_current_username(user_info)
    s = await AiQaEvalSet.create(
        project_id=pid,
        name=body.name.strip(),
        description=body.description,
        create_by=username,
    )
    return StandardResponse(data=set_to_dict(s, case_count=0), message="创建成功")


@router.put("/sets/{set_id}", response_model=StandardResponse)
async def update_set(
    set_id: int,
    body: SetBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    s = await _get_set(set_id, pid)
    s.name = body.name.strip()
    s.description = body.description
    await s.save()
    cnt = await AiQaEvalCase.filter(set_id=s.id, is_del=False).count()
    return StandardResponse(data=set_to_dict(s, case_count=cnt), message="更新成功")


@router.delete("/sets/{set_id}", response_model=StandardResponse)
async def delete_set(
    set_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    s = await _get_set(set_id, pid)
    s.is_del = True
    await s.save()
    await AiQaEvalCase.filter(set_id=set_id).update(is_del=True)
    return StandardResponse(message="已删除")


@router.get("/sets/{set_id}/cases", response_model=StandardResponse)
async def list_cases(
    set_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    cases = await list_ordered_cases(set_id)
    return StandardResponse(data=[case_to_dict(c) for c in cases])


@router.post("/sets/{set_id}/cases", response_model=StandardResponse)
async def create_case(
    set_id: int,
    body: CaseBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    ct = body.case_type if body.case_type in CASE_TYPES else "事实"
    q, points, ans = _normalize_case_input(body.question, body.expected_answer, body.expected_points)
    preset = (body.preset_answer or "").strip() or None
    extra_fields = _case_fields_from_body(body)
    c = await AiQaEvalCase.create(
        set_id=set_id,
        question=q,
        expected_points=points,
        expected_answer=ans,
        preset_answer=preset,
        category=body.category or "",
        case_type=ct,
        sort_order=body.sort_order,
        **extra_fields,
    )
    return StandardResponse(data=case_to_dict(c), message="创建成功")


@router.put("/sets/{set_id}/cases/{case_id}", response_model=StandardResponse)
async def update_case(
    set_id: int,
    case_id: int,
    body: CaseBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    c = await AiQaEvalCase.get_or_none(id=case_id, set_id=set_id, is_del=False)
    if not c:
        raise HTTPException(status_code=404, detail="用例不存在")
    q, points, ans = _normalize_case_input(body.question, body.expected_answer, body.expected_points)
    c.question = q
    c.expected_points = points
    c.expected_answer = ans
    c.preset_answer = (body.preset_answer or "").strip() or None
    c.category = body.category or ""
    c.case_type = body.case_type if body.case_type in CASE_TYPES else c.case_type
    c.sort_order = body.sort_order
    for k, v in _case_fields_from_body(body).items():
        setattr(c, k, v)
    await c.save()
    return StandardResponse(data=case_to_dict(c), message="更新成功")


@router.delete("/sets/{set_id}/cases/{case_id}", response_model=StandardResponse)
async def delete_case(
    set_id: int,
    case_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    c = await AiQaEvalCase.get_or_none(id=case_id, set_id=set_id, is_del=False)
    if not c:
        raise HTTPException(status_code=404, detail="用例不存在")
    c.is_del = True
    await c.save()
    return StandardResponse(message="已删除")


@router.post("/sets/{set_id}/cases/import", response_model=StandardResponse)
async def import_cases(
    set_id: int,
    file: UploadFile = File(...),
    project_id: Optional[int] = Query(None),
    replace: bool = Form(False),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    raw = await file.read()
    if len(raw) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=400, detail="文件超过 5MB")
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel，请使用 .xlsx 格式")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 无数据行")
    header = [str(h or "").strip() for h in rows[0]]

    def col_idx(*names: str) -> int:
        """表头列定位：先精确匹配，再「列名包含关键字」；不用短列名匹配长关键字（避免「问题」命中「问题类型」）。"""
        for n in names:
            for i, h in enumerate(header):
                if (h or "").strip() == n:
                    return i
        for n in names:
            for i, h in enumerate(header):
                hs = (h or "").strip()
                if n in hs:
                    return i
        return -1

    qi = col_idx("问题", "question", "测试问题")
    ai = col_idx("标准答案", "参考答案", "expected_answer")
    if qi < 0 or ai < 0:
        raise HTTPException(status_code=400, detail="表头需包含「问题」「标准答案」列")
    pi = col_idx("标准要点", "要点", "expected_points")
    ri = col_idx("实际回答", "模型回答", "preset_answer", "actual_answer")
    si = col_idx("序号", "seq_no", "seq")
    cpi = col_idx("问答目录", "chat_path", "chatPath")
    mi = col_idx("是否开启多轮", "是否多轮", "multi_turn", "多轮")
    sti = col_idx("问题类型", "场景类型", "scenario_type", "场景")
    fi = col_idx("文件", "source_file", "附件")
    fti = col_idx("文件类型", "file_type")
    ci = col_idx("分类", "category")
    ti = col_idx("题型", "case_type")

    if replace:
        await AiQaEvalCase.filter(set_id=set_id).update(is_del=True)

    imported = 0
    skipped = 0
    skipped_rows: list[dict[str, Any]] = []
    imported_seq_nos: list[Optional[int]] = []
    for excel_row, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        question = str(row[qi] if qi < len(row) else "").strip()
        expected_answer = str(row[ai] if ai < len(row) and row[ai] is not None else "").strip()
        seq_no_preview = _parse_int_cell(row[si] if si >= 0 and si < len(row) else None)
        if not question or not expected_answer:
            skipped += 1
            skipped_rows.append(
                {
                    "excel_row": excel_row,
                    "seq_no": seq_no_preview,
                    "reason": "缺少问题或标准答案",
                }
            )
            continue
        points = _parse_points_cell(row[pi] if pi >= 0 and pi < len(row) else None)
        preset_answer = None
        if ri >= 0 and ri < len(row) and row[ri] is not None:
            pa = str(row[ri]).strip()
            if pa:
                preset_answer = pa
        category = ""
        if ci >= 0 and ci < len(row) and row[ci]:
            category = str(row[ci]).strip()
        case_type = "事实"
        if ti >= 0 and ti < len(row) and row[ti]:
            ct = str(row[ti]).strip()
            if ct in CASE_TYPES:
                case_type = ct
        seq_no = seq_no_preview
        chat_path = _parse_chat_path_cell(row[cpi] if cpi >= 0 and cpi < len(row) else None)
        multi_turn = _parse_bool_cell(row[mi] if mi >= 0 and mi < len(row) else None)
        scenario_type = ""
        if sti >= 0 and sti < len(row) and row[sti]:
            scenario_type = str(row[sti]).strip()
        source_file = ""
        if fi >= 0 and fi < len(row) and row[fi]:
            source_file = str(row[fi]).strip()
        file_type = ""
        if fti >= 0 and fti < len(row) and row[fti]:
            file_type = str(row[fti]).strip()
        await AiQaEvalCase.create(
            set_id=set_id,
            question=question,
            expected_points=points,
            expected_answer=expected_answer,
            preset_answer=preset_answer,
            seq_no=seq_no,
            chat_path=chat_path,
            multi_turn=multi_turn,
            scenario_type=scenario_type or None,
            source_file=source_file or None,
            file_type=file_type or None,
            category=category,
            case_type=case_type,
            sort_order=imported,
        )
        imported += 1
        imported_seq_nos.append(seq_no)
    wb.close()
    seq_meta = analyze_seq_numbers(imported_seq_nos)
    missing = seq_meta.get("missing_seq_nos") or []
    dup = int(seq_meta.get("duplicate_seq_count") or 0)
    msg = f"导入 {imported} 条，跳过 {skipped} 条"
    if missing:
        preview = "、".join(str(x) for x in missing[:8])
        if len(missing) > 8:
            preview += f" 等{len(missing)}个"
        msg += f"；序号不连续，缺失：{preview}"
    if dup:
        msg += f"；重复序号 {dup} 条"
    return StandardResponse(
        data={
            "imported": imported,
            "skipped": skipped,
            "skipped_rows": skipped_rows[:30],
            "max_seq_no": seq_meta.get("max_seq_no"),
            "missing_seq_nos": missing[:50],
            "duplicate_seq_count": dup,
        },
        message=msg,
    )


@router.get("/targets", response_model=StandardResponse)
async def list_targets(
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    pid = _resolve_project_id(user_info, project_id)
    rows = await AiQaEvalTarget.filter(project_id=pid, is_del=False).order_by("-id")
    return StandardResponse(data=[target_to_dict(t) for t in rows])


@router.post("/targets", response_model=StandardResponse)
async def create_target(
    body: TargetBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    username = await resolve_current_username(user_info)
    if not (body.config.get("url") or "").strip():
        raise HTTPException(status_code=400, detail="config.url 必填")
    t = await AiQaEvalTarget.create(
        project_id=pid,
        name=body.name.strip(),
        config=body.config,
        create_by=username,
    )
    return StandardResponse(data=target_to_dict(t), message="创建成功")


@router.put("/targets/{target_id}", response_model=StandardResponse)
async def update_target(
    target_id: int,
    body: TargetBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    t = await AiQaEvalTarget.get_or_none(id=target_id, project_id=pid, is_del=False)
    if not t:
        raise HTTPException(status_code=404, detail="被测 API 配置不存在")
    if not (body.config.get("url") or "").strip():
        raise HTTPException(status_code=400, detail="config.url 必填")
    t.name = body.name.strip()
    t.config = body.config
    await t.save()
    return StandardResponse(data=target_to_dict(t), message="更新成功")


@router.delete("/targets/{target_id}", response_model=StandardResponse)
async def delete_target(
    target_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    t = await AiQaEvalTarget.get_or_none(id=target_id, project_id=pid, is_del=False)
    if not t:
        raise HTTPException(status_code=404, detail="被测 API 配置不存在")
    t.is_del = True
    await t.save()
    return StandardResponse(message="已删除")


@router.post("/targets/test", response_model=StandardResponse)
async def test_target_config(
    body: TargetTestBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    """调试被测 API 配置（保存前可用，传入 config）"""
    _resolve_project_id(user_info, project_id)
    cfg = body.config if isinstance(body.config, dict) else {}
    if not (cfg.get("url") or "").strip():
        raise HTTPException(status_code=400, detail="config.url 必填")
    data = await test_target_api(cfg, body.question.strip(), extra=body.extra)
    return StandardResponse(data=data)


@router.post("/targets/{target_id}/test", response_model=StandardResponse)
async def test_target_by_id(
    target_id: int,
    body: TargetTestBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    """调试已保存的被测 API"""
    pid = _resolve_project_id(user_info, project_id)
    t = await AiQaEvalTarget.get_or_none(id=target_id, project_id=pid, is_del=False)
    if not t:
        raise HTTPException(status_code=404, detail="被测 API 配置不存在")
    cfg = t.config if isinstance(t.config, dict) else {}
    data = await test_target_api(cfg, body.question.strip(), extra=body.extra)
    return StandardResponse(data=data)


@router.get("/targets/presets/qa-sse", response_model=StandardResponse)
async def get_qa_sse_preset(
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """问答 SSE 流式接口配置模板（不含密钥，需自行填写 Headers）"""
    from app.core.qa_eval_target_client import QA_SSE_DEFAULT_BODY, QA_SSE_PARSER_V1

    return StandardResponse(
        data={
            "name": "问答 SSE（流式）",
            "config": {
                "url": "https://kcf-pro-test.apps.digiwincloud.com.cn/api/v1/qa",
                "method": "POST",
                "response_type": "sse",
                "sse_parser": QA_SSE_PARSER_V1,
                "connect_timeout_sec": 30,
                "read_timeout_sec": 300,
                "default_body": QA_SSE_DEFAULT_BODY,
                "body_template": "",
                "answer_jsonpath": "",
                "headers": {
                    "Content-Type": "application/json",
                    "token": "",
                    "digi-middleware-auth-app": "",
                    "signature": "",
                    "signature-data": "",
                },
            },
            "header_hint": (
                "鉴权 Header 必填：token、digi-middleware-auth-app（从浏览器开发者工具复制）。"
                "若接口要求 signature，也需一并填写。"
                "default_body 中的 rootDirectory、oneDirectory、directoryType 等请按环境修改。"
            ),
        }
    )


@router.get("/question-types", response_model=StandardResponse)
async def list_question_types(
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """问题类型预设（与 RAG 评测报告一致）"""
    return StandardResponse(data=list(QA_QUESTION_TYPES))


@router.get("/import-template")
async def download_import_template(
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """下载问答评测 Excel 导入模板（含表头与示例行）"""
    rows = build_import_template_rows()
    xlsx_bytes = build_xlsx_bytes(rows, columns=QA_IMPORT_TEMPLATE_COLUMNS)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="qa_eval_template.xlsx"'
        },
    )


@router.get("/sets/{set_id}/cases/export")
async def export_set_cases(
    set_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """导出评测集用例（含已保存的预置实际回答）为 Excel"""
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    rows = await build_qa_cases_export_rows(set_id)
    xlsx_bytes = await _qa_xlsx_bytes(rows)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="qa_eval_cases.xlsx"'},
    )


@router.get("/sets/{set_id}/runs", response_model=StandardResponse)
async def list_runs(
    set_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    pid = _resolve_project_id(user_info, project_id)
    s = await _get_set(set_id, pid)
    runs = await AiQaEvalRun.filter(set_id=set_id, project_id=pid).order_by("-id").limit(50)
    targets = {t.id: t.name for t in await AiQaEvalTarget.filter(project_id=pid, is_del=False)}
    out = [run_to_dict(r, set_name=s.name, target_name=targets.get(r.target_id, "")) for r in runs]
    return StandardResponse(data=out)


@router.delete("/runs/{run_id}", response_model=StandardResponse)
async def delete_run(
    run_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    """删除跑批记录及其全部单题结果（进行中的不可删）"""
    pid = _resolve_project_id(user_info, project_id)
    try:
        await delete_eval_run(run_id, project_id=pid)
    except ValueError as e:
        msg = str(e)
        code = 409 if "进行中" in msg else 404
        raise HTTPException(status_code=code, detail=msg)
    return StandardResponse(message="已删除")


@router.get("/sets/{set_id}/runs/active", response_model=StandardResponse)
async def get_active_run(
    set_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """当前评测集进行中的跑批（用于页面恢复轮询）"""
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    run = await AiQaEvalRun.filter(
        set_id=set_id,
        project_id=pid,
        status__in=["pending", "running"],
    ).order_by("-id").first()
    if not run:
        return StandardResponse(data=None)
    s = await AiQaEvalSet.get_or_none(id=set_id)
    t = await AiQaEvalTarget.get_or_none(id=run.target_id)
    return StandardResponse(
        data=run_to_dict(
            run,
            set_name=s.name if s else "",
            target_name=t.name if t else "",
        )
    )


@router.post("/sets/{set_id}/run", response_model=StandardResponse)
async def run_eval(
    set_id: int,
    body: RunBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    s = await _get_set(set_id, pid)
    username = await resolve_current_username(user_info)
    run_mode = body.run_mode or RUN_MODE_AUTO
    target, case_count = await _validate_run_body(set_id, pid, body)
    target_id = body.target_id if run_mode != RUN_MODE_JUDGE_ONLY else body.target_id

    existing = await has_active_run(set_id, pid)
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"该评测集已有进行中的任务（#{existing.id}），请等待完成或在执行记录查看进度",
        )

    run = await AiQaEvalRun.create(
        project_id=pid,
        set_id=set_id,
        target_id=target_id,
        judge_config_id=body.judge_config_id,
        status="pending",
        create_by=username,
        extra=_build_run_extra(user_info=user_info, username=username, body=body, set_name=s.name),
    )
    asyncio.create_task(run_qa_eval_background(run.id))
    mode_label = RUN_MODE_LABELS.get(run_mode, run_mode)
    if run_mode == RUN_MODE_FETCH_ONLY:
        msg = f"批量拉取任务已提交（#{run.id}），共 {case_count} 题；完成后可下载 Excel"
    else:
        msg = f"{mode_label}任务已提交（#{run.id}），共 {case_count} 题，可关闭页面查看进度"
    return StandardResponse(
        data=run_to_dict(run, set_name=s.name, target_name=target.name if target else ""),
        message=msg,
    )


@router.post("/sets/{set_id}/runs/batch", response_model=StandardResponse)
async def run_eval_batch(
    set_id: int,
    body: BatchRunBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    """将评测集按 chunk_size 拆成多批，后台顺序执行（适合 500+ 题）"""
    pid = _resolve_project_id(user_info, project_id)
    s = await _get_set(set_id, pid)
    username = await resolve_current_username(user_info)
    if body.case_scope not in (CASE_SCOPE_ALL,):
        raise HTTPException(status_code=400, detail="自动分批仅支持「全部用例」范围")
    await _validate_run_body(set_id, pid, body)

    existing = await has_active_run(set_id, pid)
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"该评测集已有进行中的任务（#{existing.id}），请等待完成",
        )

    try:
        result = await create_batch_runs(
            set_id=set_id,
            project_id=pid,
            username=username,
            user_info=user_info,
            run_mode=body.run_mode,
            target_id=body.target_id,
            judge_config_id=body.judge_config_id,
            chunk_size=body.chunk_size,
            request_interval_ms=body.request_interval_ms,
            run_name=_normalize_run_name(body.run_name, set_name=s.name),
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    first_run = await AiQaEvalRun.get_or_none(id=result["run_ids"][0]) if result["run_ids"] else None
    t = await AiQaEvalTarget.get_or_none(id=body.target_id) if body.target_id else None
    return StandardResponse(
        data={
            **result,
            "first_run": run_to_dict(first_run, set_name=s.name, target_name=t.name if t else "")
            if first_run
            else None,
        },
        message=(
            f"已创建 {result['batch_total']} 批任务（共 {result['total_cases']} 题，"
            f"每批最多 {result['chunk_size']} 题），将顺序执行"
        ),
    )


@router.post("/sets/{set_id}/runs/merge-export")
async def merge_export_runs(
    set_id: int,
    body: MergeExportBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """合并多批跑批结果为一份 Excel（同序号后者覆盖前者）"""
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    if not body.run_ids and not body.batch_group_id:
        raise HTTPException(status_code=400, detail="请指定 run_ids 或 batch_group_id")
    rows = await build_merged_export_rows(
        run_ids=body.run_ids,
        batch_group_id=body.batch_group_id,
        set_id=set_id,
    )
    if not rows:
        raise HTTPException(status_code=400, detail="暂无结果可合并导出")
    xlsx_bytes = await _qa_xlsx_bytes(rows)
    suffix = (body.batch_group_id or "merged")[:12]
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="qa_eval_merged_{suffix}.xlsx"'},
    )


@router.post("/sets/{set_id}/runs/merge-stats-report", response_model=StandardResponse)
async def merge_stats_report(
    set_id: int,
    body: MergeExportBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """合并多批跑批结果并生成汇总统计报告"""
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    if not body.run_ids and not body.batch_group_id:
        raise HTTPException(status_code=400, detail="请指定 run_ids 或 batch_group_id")
    try:
        report = await build_merged_statistics_report(
            set_id=set_id,
            run_ids=body.run_ids,
            batch_group_id=body.batch_group_id,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return StandardResponse(data=report)


@router.post("/sets/{set_id}/runs/compare-report", response_model=StandardResponse)
async def compare_runs_report(
    set_id: int,
    body: CompareReportBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """多迭代 / 多跑批对比报告（每组可合并多条记录）"""
    pid = _resolve_project_id(user_info, project_id)
    await _get_set(set_id, pid)
    specs: list[CompareGroupSpec] = []
    for g in body.groups:
        if not g.run_ids and not g.batch_group_id:
            raise HTTPException(
                status_code=400,
                detail=f"对比组「{g.label}」需指定 run_ids 或 batch_group_id",
            )
        specs.append(
            CompareGroupSpec(
                label=g.label.strip(),
                run_ids=g.run_ids,
                batch_group_id=(g.batch_group_id or "").strip() or None,
            )
        )
    try:
        report = await build_compare_report(set_id=set_id, groups=specs)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return StandardResponse(data=report)


@router.get("/runs/{run_id}", response_model=StandardResponse)
async def get_run_report(
    run_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    pid = _resolve_project_id(user_info, project_id)
    run = await AiQaEvalRun.get_or_none(id=run_id, project_id=pid)
    if not run:
        raise HTTPException(status_code=404, detail="跑批记录不存在")
    s = await AiQaEvalSet.get_or_none(id=run.set_id)
    t = await AiQaEvalTarget.get_or_none(id=run.target_id)
    results = await AiQaEvalResult.filter(run_id=run_id).order_by("id")
    case_ids = [r.case_id for r in results]
    cases = await AiQaEvalCase.filter(id__in=case_ids) if case_ids else []
    case_map = {c.id: c for c in cases}
    failed_only = [
        result_to_dict(r, case=case_map.get(r.case_id))
        for r in results
        if not r.passed
    ]
    include_results = run.status in ("completed", "failed")
    return StandardResponse(
        data={
            "run": run_to_dict(
                run,
                set_name=s.name if s else "",
                target_name=t.name if t else "",
            ),
            "results": [
                result_to_dict(r, case=case_map.get(r.case_id)) for r in results
            ]
            if include_results
            else [],
            "failed_samples": failed_only[:20] if include_results else [],
            "partial": not include_results,
        }
    )


@router.get("/runs/{run_id}/stats-report", response_model=StandardResponse)
async def get_run_stats_report(
    run_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """统计报告（结构对齐 RAG 测试总结 docx：等级分布、场景/文件类型分析等）"""
    pid = _resolve_project_id(user_info, project_id)
    run = await AiQaEvalRun.get_or_none(id=run_id, project_id=pid)
    if not run:
        raise HTTPException(status_code=404, detail="跑批记录不存在")
    if run.status not in ("completed", "failed"):
        raise HTTPException(status_code=400, detail="任务尚未完成，请稍后再查看统计")
    try:
        report = await build_run_statistics_report(run_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return StandardResponse(data=report)


@router.post("/runs/{run_id}/results/{result_id}/regenerate", response_model=StandardResponse)
async def regenerate_run_result(
    run_id: int,
    result_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    """重新调用被测 API 并更新单条评测结果"""
    pid = _resolve_project_id(user_info, project_id)
    run = await AiQaEvalRun.get_or_none(id=run_id, project_id=pid)
    if not run:
        raise HTTPException(status_code=404, detail="跑批记录不存在")
    try:
        r = await regenerate_result(run_id, result_id, user_info=user_info)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    case = await AiQaEvalCase.get_or_none(id=r.case_id)
    run = await AiQaEvalRun.get(id=run_id)
    s = await AiQaEvalSet.get_or_none(id=run.set_id)
    t = await AiQaEvalTarget.get_or_none(id=run.target_id)
    return StandardResponse(
        data={
            "result": result_to_dict(r, case=case),
            "run": run_to_dict(run, set_name=s.name if s else "", target_name=t.name if t else ""),
        },
        message="已重新生成",
    )


@router.patch("/runs/{run_id}/results/{result_id}/review", response_model=StandardResponse)
async def review_result(
    run_id: int,
    result_id: int,
    body: ResultReviewBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    run = await AiQaEvalRun.get_or_none(id=run_id, project_id=pid)
    if not run:
        raise HTTPException(status_code=404, detail="跑批记录不存在")
    r = await AiQaEvalResult.get_or_none(id=result_id, run_id=run_id)
    if not r:
        raise HTTPException(status_code=404, detail="结果不存在")
    username = await resolve_current_username(user_info)
    from datetime import datetime

    r.manual_status = body.manual_status
    r.manual_comment = (body.manual_comment or "").strip() or None
    r.manual_reviewed_by = username
    r.manual_reviewed_at = datetime.now()
    await r.save()
    case = await AiQaEvalCase.get_or_none(id=r.case_id)
    return StandardResponse(
        data=result_to_dict(r, case=case),
        message="审核已保存",
    )


@router.post("/runs/{run_id}/results/bulk-review", response_model=StandardResponse)
async def bulk_review_results(
    run_id: int,
    body: BulkReviewBody,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_EXECUTE)),
):
    pid = _resolve_project_id(user_info, project_id)
    run = await AiQaEvalRun.get_or_none(id=run_id, project_id=pid)
    if not run:
        raise HTTPException(status_code=404, detail="跑批记录不存在")
    username = await resolve_current_username(user_info)
    from datetime import datetime

    now = datetime.now()
    updated = 0
    for rid in body.result_ids:
        r = await AiQaEvalResult.get_or_none(id=rid, run_id=run_id)
        if not r:
            continue
        r.manual_status = body.manual_status
        r.manual_comment = (body.manual_comment or "").strip() or None
        r.manual_reviewed_by = username
        r.manual_reviewed_at = now
        await r.save()
        updated += 1
    return StandardResponse(data={"updated": updated}, message=f"已更新 {updated} 条审核状态")


@router.get("/runs/{run_id}/export-answers")
async def export_run_answers(
    run_id: int,
    project_id: Optional[int] = Query(None),
    user_info: dict = Depends(require_permissions(AI_TEST_VIEW)),
):
    """导出某次跑批的问答结果 Excel（批量拉取或完整评测均可）"""
    pid = _resolve_project_id(user_info, project_id)
    run = await AiQaEvalRun.get_or_none(id=run_id, project_id=pid)
    if not run:
        raise HTTPException(status_code=404, detail="跑批记录不存在")
    if run.status not in ("completed", "failed"):
        raise HTTPException(status_code=400, detail="任务尚未完成，请稍后再导出")
    rows = await build_qa_fetch_export_rows(run_id)
    if not rows:
        raise HTTPException(status_code=400, detail="暂无结果可导出")
    xlsx_bytes = await _qa_xlsx_bytes(rows)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="qa_eval_run_{run_id}.xlsx"'},
    )
