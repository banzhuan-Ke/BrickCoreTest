"""性能测试执行记录 API"""
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, HTTPException, Query, status, Body, Depends
from fastapi.responses import HTMLResponse, Response
from pydantic import BaseModel

from app.models.perf import PerfRecord, PerfScene
from app.core.platform.auth import require_permissions, is_authenticated
from app.core.ops.notification import NotificationService
from app.core.platform.permissions import PERF_RECORD_VIEW, PERF_SCENE_EDIT
from app.core.platform.project_access import PROJECT_ROLE_MEMBER, PROJECT_ROLE_VIEWER, assert_project_access
from app.routers.perf.access import get_record_for_viewer
from app.routers.perf.report_utils import (
    build_report_payload,
    build_config_summary,
    enrich_case_aggregations,
    paginate_perf_request_items,
    render_perf_html_chart_parts,
    render_perf_html_errors,
    render_perf_html_status_codes,
    render_perf_html_stream_sections,
    render_perf_html_request_traces,
    summarize_stepping_stages,
)
from app.modules.stream_phase import detail_to_excel_row, is_stream_burst_mode, normalize_perf_mode, migrate_legacy_detail

router = APIRouter(
    prefix="/records",
    tags=["性能测试记录"],
    dependencies=[Depends(require_permissions(PERF_RECORD_VIEW))],
)

# 列表页仅需摘要字段；避免 SELECT 大 JSON 列后在 ORDER BY 时触发 sort buffer 溢出
_RECORD_LIST_FIELDS = (
    "id", "scene_id", "project_id", "status", "trigger_type",
    "config_snapshot", "total_requests", "success_count", "fail_count",
    "qps", "avg_response_time", "error_rate", "median_response_time",
    "p90_response_time", "std_dev_response_time", "received_kb_per_sec",
    "sent_kb_per_sec", "duration", "started_at", "ended_at", "run_by",
)

# 与上一条/基线对比只需指标字段；勿带 time_series / request_details 等大 JSON
_RECORD_COMPARE_FIELDS = (
    "id", "started_at", "duration", "total_requests", "success_count",
    "qps", "avg_response_time", "p95_response_time", "error_rate",
    "config_snapshot", "scene_items_snapshot", "phase_metrics",
)


async def _fetch_sibling_record_for_compare(record_id: int) -> Optional[PerfRecord]:
    if not record_id:
        return None
    return await PerfRecord.filter(id=record_id).only(*_RECORD_COMPARE_FIELDS).first()


async def _fetch_previous_record_for_compare(record: PerfRecord) -> Optional[PerfRecord]:
    """先只查 id 再按主键取对比字段，避免 ORDER BY 全行大 JSON 触发 MySQL sort buffer 溢出。"""
    prev = await (
        PerfRecord.filter(
            scene_id=record.scene_id,
            id__lt=record.id,
            status__in=["success", "failed", "stopped"],
        )
        .order_by("-id")
        .only("id")
        .first()
    )
    if not prev:
        return None
    return await _fetch_sibling_record_for_compare(prev.id)



@router.get("", summary="执行记录列表")
async def get_records(
    project_id: int = Query(..., description="项目ID"),
    status: Optional[str] = Query(None, description="状态筛选"),
    scene_name: Optional[str] = Query(None, description="场景名称关键字"),
    start_date: Optional[str] = Query(None, description="开始日期(YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期(YYYY-MM-DD)"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=5000),
    user_info: dict = Depends(is_authenticated),
):
    """获取性能测试执行记录列表（支持按场景名称、状态、时间范围筛选）"""
    await assert_project_access(user_info, project_id, min_role=PROJECT_ROLE_VIEWER)
    query = PerfRecord.filter(project_id=project_id)

    if status:
        query = query.filter(status=status)

    if scene_name:
        # 通过 scene_id 关联查询场景名称
        scenes = await PerfScene.filter(project_id=project_id, name__contains=scene_name, is_del=False).all()
        scene_ids = [s.id for s in scenes]
        if scene_ids:
            query = query.filter(scene_id__in=scene_ids)
        else:
            # 无匹配场景，返回空结果
            return {"data": [], "total": 0, "page": page, "size": size}

    if start_date:
        query = query.filter(started_at__gte=start_date)
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(started_at__lte=end_dt)
        except ValueError:
            query = query.filter(started_at__lte=end_date)

    total = await query.count()
    records = await (
        query.order_by("-id")
        .offset((page - 1) * size)
        .limit(size)
        .only(*_RECORD_LIST_FIELDS)
        .prefetch_related("scene")
        .all()
    )

    result = []
    for record in records:
        scene = record.scene
        result.append({
            "id": record.id,
            "scene_id": record.scene_id,
            "scene_name": scene.name if scene else "未知场景",
            "project_id": record.project_id,
            "status": record.status,
            "trigger_type": record.trigger_type,
            "config_snapshot": record.config_snapshot,
            "mode": record.config_snapshot.get("mode", "fixed") if record.config_snapshot else "fixed",
            "total_requests": record.total_requests,
            "success_count": record.success_count,
            "fail_count": record.fail_count,
            "qps": record.qps,
            "avg_response_time": record.avg_response_time,
            "error_rate": record.error_rate,
            "median_response_time": record.median_response_time,
            "p90_response_time": record.p90_response_time,
            "std_dev_response_time": record.std_dev_response_time,
            "received_kb_per_sec": record.received_kb_per_sec,
            "sent_kb_per_sec": record.sent_kb_per_sec,
            "duration": record.duration,
            "started_at": record.started_at.strftime("%Y-%m-%d %H:%M:%S") if record.started_at else None,
            "ended_at": record.ended_at.strftime("%Y-%m-%d %H:%M:%S") if record.ended_at else None,
            "run_by": record.run_by
        })
    
    return {"data": result, "total": total, "page": page, "size": size}


@router.get("/{record_id}", summary="执行记录详情")
async def get_record_detail(record: PerfRecord = Depends(get_record_for_viewer)):
    """获取性能测试执行记录详情"""
    scene = await record.scene
    try:
        previous = await _fetch_previous_record_for_compare(record)
    except Exception:
        previous = None

    baseline_record = None
    pinned_id = getattr(scene, "baseline_record_id", None) if scene else None
    if pinned_id and pinned_id != record.id:
        try:
            baseline_record = await _fetch_sibling_record_for_compare(pinned_id)
        except Exception:
            baseline_record = None

    return await build_report_payload(record, scene, previous, baseline_record=baseline_record)


@router.get("/{record_id}/request-items", summary="分页获取请求明细（报告懒加载）")
async def get_request_items(
    record: PerfRecord = Depends(get_record_for_viewer),
    kind: str = Query("stream", description="stream=流式阶段明细, http=HTTP请求trace"),
    status: str = Query("all", description="all|success|fail"),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=100),
):
    if kind not in ("stream", "http"):
        raise HTTPException(status_code=400, detail="kind 须为 stream 或 http")
    return paginate_perf_request_items(record, kind=kind, status=status, page=page, size=size)


@router.post(
    "/batch-delete",
    summary="批量删除执行记录",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(require_permissions(PERF_SCENE_EDIT))],
)
async def batch_delete_records(
    record_ids: List[int],
    user_info: dict = Depends(is_authenticated),
):
    """批量删除性能测试执行记录（先全量鉴权，再事务删除）。"""
    if not record_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的记录")

    # 去重并保持稳定顺序
    unique_ids = list(dict.fromkeys(record_ids))
    records = await PerfRecord.filter(id__in=unique_ids).all()
    found = {r.id: r for r in records}
    missing = [rid for rid in unique_ids if rid not in found]
    if missing:
        raise HTTPException(status_code=404, detail=f"记录不存在: {missing[:10]}")

    # 先校验全部权限，避免半成功删除
    for record in records:
        await assert_project_access(user_info, record.project_id, min_role=PROJECT_ROLE_MEMBER)

    from tortoise.transactions import in_transaction

    async with in_transaction():
        await PerfRecord.filter(id__in=unique_ids).delete()


class SendPerfReportPayload(BaseModel):
    config_ids: Optional[List[int]] = None
    recipients: Optional[List[str]] = None


@router.post("/{record_id}/send-report", summary="发送性能测试报告", status_code=status.HTTP_200_OK)
async def send_perf_report(
    record: PerfRecord = Depends(get_record_for_viewer),
    payload: Optional[SendPerfReportPayload] = None,
):
    """手动发送性能测试报告（可指定通知配置 id）"""
    body = payload or SendPerfReportPayload()
    try:
        result = await NotificationService.send_perf_report(
            project_id=record.project_id,
            record_id=record.id,
            recipients=body.recipients,
            config_ids=body.config_ids,
        )
        return {"detail": NotificationService.format_dispatch_detail(result)}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"发送失败: {e}")


@router.get("/{record_id}/export", summary="导出性能测试报告(HTML)")
async def export_perf_report(record: PerfRecord = Depends(get_record_for_viewer)):
    """导出性能测试报告为 HTML"""
    from app.core.shared.download_headers import content_disposition_attachment, sanitize_download_basename

    scene = await record.scene
    html_content = _generate_perf_html_report(record, scene)
    scene_name = sanitize_download_basename(
        getattr(scene, "name", None) or "性能测试报告",
        fallback="性能测试报告",
    )
    filename = f"{scene_name}.html"
    ascii_fallback = f"perf_report_{record.id}.html"
    return HTMLResponse(
        content=html_content,
        headers=content_disposition_attachment(filename, ascii_fallback=ascii_fallback),
    )


@router.get("/{record_id}/export-excel", summary="导出SSE阶段压测Excel报告")
async def export_perf_excel(record: PerfRecord = Depends(get_record_for_viewer)):
    """导出 SSE 问答阶段压测明细与汇总 Excel（双 sheet）。"""
    mode = normalize_perf_mode((record.config_snapshot or {}).get("mode"))
    details = [migrate_legacy_detail(d) for d in (record.request_details or [])]
    if not details or not is_stream_burst_mode(mode):
        raise HTTPException(status_code=400, detail="该记录无流式阶段压测明细，请使用流式阶段压测模式执行")

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    import io

    wb = Workbook()
    detail_ws = wb.active
    detail_ws.title = "详细结果"

    detail_rows = [detail_to_excel_row(i, d) for i, d in enumerate(details, 1)]
    if detail_rows:
        headers = list(detail_rows[0].keys())
        detail_ws.append(headers)
        for row in detail_rows:
            detail_ws.append([row.get(h, "") for h in headers])
        for col_idx, _ in enumerate(headers, 1):
            detail_ws.column_dimensions[get_column_letter(col_idx)].width = min(30, 16)

    summary_ws = wb.create_sheet("汇总")
    phase_metrics = record.phase_metrics or {}
    metrics = phase_metrics.get("metrics") or []
    if metrics:
        summary_headers = ["指标", "并发数", "正常请求数", "异常量", "异常率(%)", "平均数(s)", "中位数(s)", "90%百分位(s)", "95%百分位(s)", "99%百分位(s)", "最小值(s)", "最大值(s)"]
        summary_ws.append(summary_headers)
        for m in metrics:
            summary_ws.append([
                m.get("label", ""),
                m.get("total_requests", ""),
                m.get("normal_count", ""),
                m.get("abnormal_count", ""),
                f"{m.get('error_rate', 0):.2f}%",
                m.get("mean", "N/A"),
                m.get("median", "N/A"),
                m.get("p90", "N/A"),
                m.get("p95", "N/A"),
                m.get("p99", "N/A"),
                m.get("min", "N/A"),
                m.get("max", "N/A"),
            ])
        for col_idx in range(1, len(summary_headers) + 1):
            summary_ws.column_dimensions[get_column_letter(col_idx)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    from app.core.shared.download_headers import content_disposition_attachment, sanitize_download_basename
    scene = await record.scene
    scene_name = sanitize_download_basename(
        getattr(scene, "name", None) or "流式阶段报告",
        fallback="流式阶段报告",
    )
    filename = f"{scene_name}-阶段明细.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=content_disposition_attachment(filename, ascii_fallback=f"perf_phase_{record.id}.xlsx"),
    )


@router.get("/{record_id}/report", summary="执行报告数据")
async def get_record_report(record: PerfRecord = Depends(get_record_for_viewer)):
    """获取性能测试报告数据（用于前端 echarts）"""
    scene = await record.scene
    try:
        previous = await _fetch_previous_record_for_compare(record)
    except Exception:
        previous = None

    baseline_record = None
    pinned_id = getattr(scene, "baseline_record_id", None) if scene else None
    if pinned_id and pinned_id != record.id:
        try:
            baseline_record = await _fetch_sibling_record_for_compare(pinned_id)
        except Exception:
            baseline_record = None

    return await build_report_payload(record, scene, previous, baseline_record=baseline_record)


def _generate_perf_html_report(
    record: PerfRecord,
    scene,
    *,
    editable: bool = True,
    embed_echarts: bool = True,
) -> str:
    """生成性能测试 HTML 报告（汇报型模块结构）。

    embed_echarts=False：图表库走 CDN，适合邮件附件（避免内嵌约 1MB JS）。
    """
    from app.modules.perf.perf_html_theme import (
        case_note_map,
        err_tag,
        h,
        render_conclusion_box,
        render_overview_para,
        render_percentile_table,
        render_rt_bars,
        report_css,
        report_edit_chrome,
        render_metric_glossary,
        render_target_evaluation_section,
        target_card_note,
        ai_is_done,
    )
    from app.modules.perf.perf_target_eval import evaluate_perf_targets

    scene_name = scene.name if scene else "未知场景"
    config = record.config_snapshot or {}
    cfg_summary = build_config_summary(config, record.distribution_info or {})
    mode_label = cfg_summary.get("mode_label", config.get("mode", "fixed"))
    dist_label = cfg_summary.get("distribution_mode_label", "随机权重")
    concurrent_label = cfg_summary.get("concurrent_users_label") or "并发用户数"
    concurrent_display = cfg_summary.get("concurrent_users_display")
    if concurrent_display is None:
        concurrent_display = config.get("concurrent_users", "-")
    stepping_stages = (
        summarize_stepping_stages(config, record.time_series_data or [])
        if cfg_summary.get("mode") == "stepping"
        else []
    )
    stage_rows = []
    for st in stepping_stages:
        stage_rows.append(
            "<tr>"
            f"<td>第 {h(st.get('stage'))} 阶段</td>"
            f"<td>{h(st.get('users'))}</td>"
            f"<td>{h(st.get('planned_duration') if st.get('planned_duration') is not None else '-')}</td>"
            f"<td>{h(st.get('observed_seconds') if st.get('observed_seconds') is not None else '-')}</td>"
            f"<td>{h(st.get('completed_seconds') if st.get('completed_seconds') is not None else '-')}</td>"
            f"<td>{h(st.get('avg_qps') if st.get('avg_qps') is not None else '-')}</td>"
            f"<td>{h(st.get('avg_rt') if st.get('avg_rt') is not None else '-')}"
            f"{('（约 ' + str(round(float(st['avg_rt'])/1000, 1)) + 's）') if st.get('avg_rt') is not None else ''}</td>"
            f"<td>{h(st.get('avg_p95') if st.get('avg_p95') is not None else '-')}"
            f"{('（约 ' + str(round(float(st['avg_p95'])/1000, 1)) + 's）') if st.get('avg_p95') is not None else ''}</td>"
            f"<td>{h(st.get('avg_error_rate') if st.get('avg_error_rate') is not None else '-')}</td>"
            "</tr>"
        )
    stages_html = ""
    if stage_rows:
        stages_html = f"""
    <h3 style="margin-top:18px">梯度阶段明细</h3>
    <p style="font-size:12px;color:#64748b;margin-bottom:8px">
      按配置阶段列出计划并发/时长。平均 QPS 含整段墙钟；平均 RT/P95 仅统计有完成请求的秒，避免空闲秒把耗时稀释。
    </p>
    <div class="stage-summary-block" style="margin-bottom:12px">
      <div class="stage-summary-label">分阶段摘要</div>
      <div class="stage-summary-list">
        {''.join(
          f"<div class='stage-summary-card flat'>"
          f"<div class='stage-sum-title'><span class='stage-idx'>{h(st.get('stage'))}</span>"
          f"<span>第 {h(st.get('stage'))} 阶段 · {h(st.get('users'))} 并发</span></div>"
          f"<div class='stage-sum-meta'>有完成 {h(st.get('completed_seconds') if st.get('completed_seconds') is not None else '-')}s"
          f" · QPS {h(st.get('avg_qps') if st.get('avg_qps') is not None else '-')}"
          f" · RT {h(st.get('avg_rt') if st.get('avg_rt') is not None else '-')}ms"
          f"{('（约 ' + str(round(float(st['avg_rt'])/1000, 1)) + 's）') if st.get('avg_rt') is not None else ''}"
          f" · P95 {h(st.get('avg_p95') if st.get('avg_p95') is not None else '-')}ms</div>"
          f"</div>"
          for st in stepping_stages
        )}
      </div>
    </div>
    <table>
      <thead><tr>
        <th>阶段</th><th>并发</th><th>计划时长(s)</th><th>观察秒数</th><th>有完成秒数</th>
        <th>平均 QPS</th><th>平均 RT(ms)</th><th>平均 P95(ms)</th><th>平均错误率(%)</th>
      </tr></thead>
      <tbody>{''.join(stage_rows)}</tbody>
    </table>"""

    status_class = (
        "status-success" if record.status == "success"
        else "status-fail" if record.status == "failed"
        else "status-running"
    )
    status_text = {
        "success": "成功", "failed": "失败", "running": "执行中",
        "stopped": "已停止", "pending": "等待中",
    }.get(record.status, record.status)

    generate_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    started = record.started_at.strftime("%Y-%m-%d %H:%M:%S") if record.started_at else "-"
    ended = record.ended_at.strftime("%Y-%m-%d %H:%M:%S") if record.ended_at else "-"

    ai = record.ai_analysis if isinstance(record.ai_analysis, dict) else {}
    notes_by_case = case_note_map(ai)
    target_evaluation = evaluate_perf_targets(record)

    # case_id → 最严重目标判定
    _st_rank = {"fail": 3, "warn": 2, "unknown": 1, "pass": 0, "skipped": -1}
    _st_label = {"pass": "通过", "warn": "警告", "fail": "失败", "skipped": "跳过", "unknown": "未判定"}
    case_target_best: dict[str, str] = {}
    for it in target_evaluation.get("items") or []:
        if not isinstance(it, dict) or it.get("scope") != "case":
            continue
        cid = it.get("case_id")
        if cid is None:
            continue
        st = str(it.get("status") or "unknown")
        key = str(cid)
        prev = case_target_best.get(key)
        if prev is None or _st_rank.get(st, 0) > _st_rank.get(prev, 0):
            case_target_best[key] = st
    show_case_target_col = bool(case_target_best)

    case_aggs = enrich_case_aggregations(record.case_aggregations, record.duration or 0)
    case_rows = []
    for _cid, agg in case_aggs.items():
        name = agg.get("name") or "未知"
        cnote = notes_by_case.get(name, "")
        note_html = f'<div class="case-note">{h(cnote)}</div>' if cnote else ""
        err_rate = agg.get("error_rate", 0) or 0
        target_cell = ""
        if show_case_target_col:
            st = case_target_best.get(str(_cid), "")
            if st:
                tag_cls = {
                    "pass": "tag-green", "warn": "tag-orange", "fail": "tag-red",
                    "skipped": "tag-blue", "unknown": "tag-blue",
                }.get(st, "tag-blue")
                target_cell = f'<td><span class="tag {tag_cls}">{h(_st_label.get(st, st))}</span></td>'
            else:
                target_cell = "<td>—</td>"
        case_rows.append(
            f"<tr><td>{h(name)}{note_html}</td>"
            f"<td>{h(agg.get('total', 0))}</td>"
            f"<td>{h(agg.get('fail', 0))}</td>"
            f"<td>{err_tag(err_rate)}</td>"
            f"<td>{h(agg.get('avg_rt', 0))} <span class='unit'>ms</span></td>"
            f"<td>{h(agg.get('median_rt', 0))} <span class='unit'>ms</span></td>"
            f"<td>{h(agg.get('p90_rt', 0))} <span class='unit'>ms</span></td>"
            f"<td>{h(agg.get('p95_rt', 0))} <span class='unit'>ms</span></td>"
            f"<td>{h(agg.get('max_rt', 0))} <span class='unit'>ms</span></td>"
            f"{target_cell}</tr>"
        )
    case_colspan = 10 if show_case_target_col else 9
    case_body = "".join(case_rows) or (
        f'<tr><td colspan="{case_colspan}" style="text-align:center;color:#999;">无接口数据</td></tr>'
    )
    case_target_th = "<th>目标判定</th>" if show_case_target_col else ""

    chart_sections, chart_scripts = render_perf_html_chart_parts(
        record,
        ai_trend_note=str((ai or {}).get("trend_note") or "").strip(),
        ai_dist_note=str((ai or {}).get("distribution_note") or "").strip(),
        embed_echarts=embed_echarts,
    )
    stream_html = render_perf_html_stream_sections(record)
    status_code_html = render_perf_html_status_codes(record)
    error_html = render_perf_html_errors(record)
    request_traces_html = render_perf_html_request_traces(record)

    # 错误 / 状态码 / 趋势 / 结论章节号顺延，避免双「五、」
    _cn = "零一二三四五六七八九十"

    def _sec(n: int, title: str) -> str:
        prefix = _cn[n] if 0 <= n < len(_cn) else str(n)
        return f"{prefix}、{title}"

    sec_n = 5
    if status_code_html:
        status_code_html = status_code_html.replace(
            "<h2>HTTP 状态码分布</h2>", f"<h2>{_sec(sec_n, 'HTTP 状态码分布')}</h2>", 1
        )
        sec_n += 1
    if error_html:
        error_html = error_html.replace(
            "<h2>错误分类统计</h2>", f"<h2>{_sec(sec_n, '错误分类统计')}</h2>", 1
        )
        sec_n += 1

    qps_note, qps_cls = target_card_note(
        target_evaluation, target_key="qps", ai=ai, ai_note_key="qps"
    )
    avg_note, avg_cls = target_card_note(
        target_evaluation, target_key="avg_response_time", ai=ai, ai_note_key="avg_rt"
    )
    p95_note, p95_cls = target_card_note(
        target_evaluation, target_key="p95_response_time", ai=ai, ai_note_key="p95"
    )
    err_note, err_cls = target_card_note(
        target_evaluation, target_key="error_rate", ai=ai, ai_note_key="error_rate"
    )
    total_note, total_cls = target_card_note(
        target_evaluation, target_key="total_requests", ai=ai, ai_note_key="total_requests"
    )
    # 未配目标时仍保留粗样式兜底
    if not p95_cls and (record.p95_response_time or 0) > 500:
        p95_cls = "warning"
    if not err_cls:
        err_cls = "danger" if (record.error_rate or 0) > 5 else "success"

    target_section_html = render_target_evaluation_section(
        target_evaluation, heading="二（附）、性能目标明细"
    )

    if ai_is_done(ai) and ai.get("summary"):
        conclusion = render_conclusion_box(ai)
    elif ai.get("status") in ("running", "pending"):
        conclusion = '<div class="conclusion-box warn"><p>AI 分析进行中，请稍后重新导出或在报告页刷新查看。</p></div>'
    else:
        conclusion = (
            f'<div class="conclusion-box"><p><strong>规则摘要（未跑 AI）：</strong>'
            f'QPS {round(record.qps or 0, 2)}，P95 {round(record.p95_response_time or 0, 2)} ms，'
            f'错误率 {round(record.error_rate or 0, 2)}%，总请求 {record.total_requests or 0}。</p></div>'
        )

    pct_table = render_percentile_table(
        min_rt=record.min_response_time,
        median_rt=record.median_response_time,
        avg_rt=record.avg_response_time,
        p90_rt=record.p90_response_time,
        p95_rt=record.p95_response_time,
        p99_rt=record.p99_response_time,
        max_rt=record.max_response_time,
    )
    bars = render_rt_bars(
        min_rt=record.min_response_time,
        median_rt=record.median_response_time,
        avg_rt=record.avg_response_time,
        p90_rt=record.p90_response_time,
        p95_rt=record.p95_response_time,
        max_rt=record.max_response_time,
    )

    appendix_parts = []
    if stream_html:
        appendix_parts.append(stream_html)
    if request_traces_html:
        appendix_parts.append(request_traces_html)
    appendix = ""
    if appendix_parts:
        appendix = f"""
  <details class="section-fold">
    <summary>附录：诊断与采样明细（点击展开）</summary>
    {''.join(appendix_parts)}
  </details>"""

    err_section = status_code_html + error_html
    if not err_section.strip():
        err_section = f"""
  <div class="section">
    <h2>{_sec(sec_n, '错误与状态码')}</h2>
    <p style="color:#666;font-size:13px">本次无错误分类 / 状态码分布数据。</p>
  </div>"""
        sec_n += 1

    if chart_sections:
        chart_sections = chart_sections.replace(
            "<h2>性能趋势</h2>", f"<h2>{_sec(sec_n, '性能趋势')}</h2>", 1
        )
        sec_n += 1
        _eb = record.error_breakdown if isinstance(getattr(record, "error_breakdown", None), dict) else {}
        if (_eb or {}).get("rt_histogram"):
            chart_sections = chart_sections.replace(
                "<h2>响应时间分布</h2>", f"<h2>{_sec(sec_n, '响应时间分布')}</h2>", 1
            )
            sec_n += 1
    conclusion_h2 = _sec(sec_n, "结论与建议")

    edit_chrome = report_edit_chrome() if editable else ""
    root_ce = ' contenteditable="true"' if editable else ""

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>性能测试报告 - {h(scene_name)}</title>
{report_css()}
</head>
<body>
{edit_chrome}
<div class="container" id="reportRoot"{root_ce}>
  <div class="header">
    <h1>性能测试报告 — {h(scene_name)}</h1>
    <div class="meta">
      <span class="status-badge {status_class}">{h(status_text)}</span>
      <span>模式：{h(mode_label)}</span>
      <span>分配：{h(dist_label)}</span>
      <span>执行人：{h(record.run_by or '-')}</span>
      <span>生成：{h(generate_time)}</span>
    </div>
    <div class="meta" style="margin-top:8px">
      <span>开始：{h(started)}</span>
      <span>结束：{h(ended)}</span>
      <span>时长：{h(round(record.duration or 0, 2))}s</span>
      <span>目标：{h(config.get('target_host') or '默认')}</span>
    </div>
  </div>

  <div class="section">
    <h2>一、测试概览</h2>
    {render_overview_para(ai)}
    <table class="config-table">
      <tr><td>压测模式</td><td>{h(mode_label)}</td></tr>
      <tr><td>{h(concurrent_label)}</td><td>{h(concurrent_display)}</td></tr>
      <tr><td>Ramp-up</td><td>{h(config.get('ramp_up_seconds', 0))}s</td></tr>
      <tr><td>持续时间/循环</td><td>{h(cfg_summary.get('duration_label', '-'))}</td></tr>
      <tr><td>分配模式</td><td>{h(dist_label)}</td></tr>
      <tr><td>预热</td><td>{h(cfg_summary.get('warmup_seconds') if cfg_summary.get('warmup_seconds') is not None else (config.get('warmup_seconds') or 0))}s</td></tr>
      <tr><td>总请求 / 成功 / 失败</td><td>{h(record.total_requests)} / {h(record.success_count)} / {h(record.fail_count)}</td></tr>
      <tr><td>网络吞吐</td><td>收 {h(round(record.received_kb_per_sec or 0, 2))} KB/s · 发 {h(round(record.sent_kb_per_sec or 0, 2))} KB/s</td></tr>
    </table>
    {stages_html}
  </div>

  <div class="section">
    <h2>二、核心指标</h2>
    <div class="summary-grid">
      <div class="summary-card">
        <div class="label">QPS</div>
        <div class="value {qps_cls}">{h(round(record.qps or 0, 2))} <span class="unit">/s</span></div>
        {f'<div class="note">{h(qps_note)}</div>' if qps_note else ''}
      </div>
      <div class="summary-card">
        <div class="label">平均响应时间</div>
        <div class="value {avg_cls}">{h(round(record.avg_response_time or 0, 2))} <span style="font-size:14px">ms</span></div>
        {f'<div class="note">{h(avg_note)}</div>' if avg_note else ''}
      </div>
      <div class="summary-card">
        <div class="label">P95 响应时间</div>
        <div class="value {p95_cls}">{h(round(record.p95_response_time or 0, 2))} <span style="font-size:14px">ms</span></div>
        {f'<div class="note">{h(p95_note)}</div>' if p95_note else ''}
      </div>
      <div class="summary-card">
        <div class="label">错误率</div>
        <div class="value {err_cls}">{h(round(record.error_rate or 0, 2))}%</div>
        {f'<div class="note">{h(err_note)}</div>' if err_note else ''}
      </div>
      <div class="summary-card">
        <div class="label">总请求数</div>
        <div class="value {total_cls}">{h(record.total_requests or 0)} <span class="unit">次</span></div>
        {f'<div class="note">{h(total_note)}</div>' if total_note else ''}
      </div>
    </div>
  </div>

  {target_section_html}

  <div class="section">
    <h2>三、响应时间分布</h2>
    <h3>3.1 百分位</h3>
    {pct_table}
    <h3>3.2 可视化</h3>
    {bars or '<p style="color:#999;font-size:13px">暂无分位数据</p>'}
  </div>

  <div class="section">
    <h2>四、接口明细</h2>
    <table>
      <thead>
        <tr><th>接口</th><th>请求数</th><th>失败</th><th>失败率</th>
          <th>Avg <span class="unit">(ms)</span></th>
          <th>P50 <span class="unit">(ms)</span></th>
          <th>P90 <span class="unit">(ms)</span></th>
          <th>P95 <span class="unit">(ms)</span></th>
          <th>Max <span class="unit">(ms)</span></th>{case_target_th}</tr>
      </thead>
      <tbody>{case_body}</tbody>
    </table>
  </div>

  {err_section}

  {chart_sections}

  <div class="section">
    <h2>{conclusion_h2}</h2>
    {conclusion}
  </div>

  {appendix}

  {render_metric_glossary()}

  <div class="footer">BrickCore 性能测试报告 — {h(generate_time)}</div>
</div>
{chart_scripts}
</body>
</html>"""


class CompareRequest(BaseModel):
    record_ids: List[int]
    reference_record_id: Optional[int] = None


@router.post("/compare", summary="报告对比/合订预览")
async def compare_records(
    req: CompareRequest,
    user_info: dict = Depends(is_authenticated),
):
    """选择 2–10 条执行记录：同场景对比矩阵，跨场景合订预览（不落库）。"""
    from app.modules.perf.compare_report import (
        MAX_COMPARE_RECORDS,
        MIN_COMPARE_RECORDS,
        build_report_snapshot,
    )

    record_ids = list(dict.fromkeys(req.record_ids or []))
    if len(record_ids) < MIN_COMPARE_RECORDS:
        raise HTTPException(status_code=400, detail=f"请至少选择 {MIN_COMPARE_RECORDS} 条记录")
    if len(record_ids) > MAX_COMPARE_RECORDS:
        raise HTTPException(status_code=400, detail=f"最多支持 {MAX_COMPARE_RECORDS} 条记录")

    records_map = {r.id: r for r in await PerfRecord.filter(id__in=record_ids).all()}
    missing = [i for i in record_ids if i not in records_map]
    if missing:
        raise HTTPException(status_code=404, detail=f"记录不存在: {missing[:10]}")
    records = [records_map[i] for i in record_ids]

    for r in records:
        await assert_project_access(user_info, r.project_id, min_role=PROJECT_ROLE_VIEWER)

    scene_names: dict = {}
    for r in records:
        scene = await r.scene
        scene_names[r.id] = scene.name if scene else "未知场景"

    return build_report_snapshot(
        records,
        reference_record_id=req.reference_record_id,
        scene_names=scene_names,
    )
